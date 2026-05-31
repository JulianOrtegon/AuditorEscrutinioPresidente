"""
Auditor Escrutinio Presidencial 2026
Backend Flask minimo: sesion + login + healthcheck.
Módulos se van migrando del Auditor Congreso uno por uno.
"""
from flask import Flask, jsonify, send_from_directory, request, session
import psycopg
from psycopg.rows import dict_row
from psycopg_pool import ConnectionPool
import os
import hashlib
import time
import sys
import logging

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

logging.basicConfig(level=logging.INFO, stream=sys.stderr,
                    format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger('presidencial')

app = Flask(__name__, static_folder='public')
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'tyse-escrutinio-presidencial-2026-secretkey')
app.config['MAX_CONTENT_LENGTH'] = 4 * 1024 * 1024 * 1024
app.config['MAX_FORM_MEMORY_SIZE'] = 100 * 1024 * 1024
app.config['MAX_FORM_PARTS'] = 1000

# ==================== BD ====================
DB_CONFIG = {
    'host':     os.environ.get('DB_HOST',     '192.168.0.54'),
    'port':     int(os.environ.get('DB_PORT', 5432)),
    'user':     os.environ.get('DB_USER',     'postgres'),
    'password': os.environ.get('DB_PASSWORD', 'postgres'),
    'dbname':   os.environ.get('DB_NAME',     'AuditorEscrutinioPresidencial2026_PROD'),
}

_conninfo = psycopg.conninfo.make_conninfo(**DB_CONFIG)
def _configure_conn(conn):
    conn.row_factory = dict_row

db_pool = ConnectionPool(_conninfo, min_size=5, max_size=32, open=False, configure=_configure_conn)
try:
    db_pool.open()
    logger.info(f"[bd] Pool conectado a {DB_CONFIG['host']}/{DB_CONFIG['dbname']}")
except Exception as e:
    logger.error(f"[bd] No se pudo abrir pool: {e}")

def get_db_connection():
    return db_pool.connection()

def hash_password(p):
    return hashlib.sha256(p.encode('utf-8')).hexdigest()

# ==================== MIDDLEWARE ====================
@app.before_request
def log_y_actividad():
    if request.path.startswith('/api/') and 'task-status' not in request.path:
        logger.info(f"[REQ] {request.method} {request.path} from={request.remote_addr} user={session.get('user_id','?')}")
    if 'user_id' in session:
        session['last_activity'] = time.time()

# Perfil restringido: solo puede ver y descargar el Generador Incremental.
# El resto de endpoints /api/ le devuelven 403. (Operar el generador —subir
# plantilla, iniciar, detener, ejecutar— ya está bloqueado aparte por _is_admin.)
PERFIL_SOLO_GENERADOR = 'Generador Incremental'
_GEN_ALLOW_EXACT = {'/api/session', '/api/logout', '/api/login', '/api/dashboard/metricas'}

@app.before_request
def _restringir_perfil_solo_generador():
    if not request.path.startswith('/api/'):
        return  # SPA, estáticos y pantalla de login quedan libres
    if session.get('perfil') != PERFIL_SOLO_GENERADOR:
        return  # otros perfiles: comportamiento sin cambios
    p = request.path
    if p in _GEN_ALLOW_EXACT or p.startswith('/api/generador'):
        return
    return jsonify({'success': False,
                    'error': 'Tu perfil solo tiene acceso al Generador Incremental'}), 403

@app.after_request
def no_cache(response):
    if response.content_type and ('text/html' in response.content_type or 'application/javascript' in response.content_type):
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# ==================== STATIC ====================
@app.route('/')
def index():
    return send_from_directory('public', 'index.html')

@app.route('/<path:filename>')
def static_files(filename):
    return send_from_directory('public', filename)

# ==================== AUTH ====================
@app.route('/api/login', methods=['POST'])
def login():
    try:
        data = request.get_json() or {}
        cedula = data.get('cedula', '').strip()
        contrasena = data.get('contrasena', '')
        if not cedula or not contrasena:
            return jsonify({'success': False, 'error': 'Cédula y contraseña son requeridos'}), 400
        hashed = hash_password(contrasena)
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('''
                    SELECT u.id, u.cedula, u.nombres, u.apellidos, u.correo,
                           u.id_perfil, p.nombre AS perfil_nombre
                    FROM usuarios u
                    LEFT JOIN perfiles p ON p.id = u.id_perfil
                    WHERE u.cedula = %s AND u.contrasena = %s
                ''', (cedula, hashed))
                user = cur.fetchone()
        if not user:
            return jsonify({'success': False, 'error': 'Cédula o contraseña incorrectos'}), 401

        session['user_id']   = user['id']
        session['cedula']    = user['cedula']
        session['nombres']   = user['nombres']
        session['apellidos'] = user['apellidos']
        session['perfil']    = user['perfil_nombre']
        session['id_perfil'] = user['id_perfil']
        session['last_activity'] = time.time()
        return jsonify({'success': True, 'data': {
            'id': user['id'], 'cedula': user['cedula'],
            'nombres': user['nombres'], 'apellidos': user['apellidos'],
            'perfil': user['perfil_nombre'], 'id_perfil': user['id_perfil']
        }})
    except Exception as e:
        logger.exception("[login] error")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/session', methods=['GET'])
def get_session():
    if 'user_id' not in session:
        return jsonify({'success': False, 'session_expired': True}), 401
    return jsonify({'success': True, 'data': {
        'id': session['user_id'], 'cedula': session.get('cedula'),
        'nombres': session.get('nombres'), 'apellidos': session.get('apellidos'),
        'perfil': session.get('perfil'), 'id_perfil': session.get('id_perfil')
    }})

# ==================== DIVIPOL ====================
def _require_session():
    if 'user_id' not in session:
        return jsonify({'success': False, 'session_expired': True}), 401
    return None

@app.route('/api/divipol', methods=['GET'])
def divipol_listar():
    """Listado adaptativo según filtros (mismo shape que Congreso)."""
    err = _require_session()
    if err: return err
    coddepto = request.args.get('coddepto', type=int)
    codmipio = request.args.get('codmipio', type=int)
    codzona  = request.args.get('codzona', type=int)
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                if codzona is not None:
                    cur.execute('''
                        SELECT codpuesto, nompuesto,
                               COALESCE(potmasculino,0) AS hombre,
                               COALESCE(potfemenino,0)  AS mujeres,
                               COALESCE(pottotal,0)     AS total_potencial,
                               COALESCE(nummesas,0)     AS mesas
                        FROM divipol_presidencial_2026
                        WHERE clase='P' AND coddepto=%s AND codmipio=%s AND codzona=%s
                        ORDER BY codpuesto
                    ''', (coddepto, codmipio, codzona))
                elif codmipio is not None:
                    cur.execute('''
                        SELECT codzona,
                               COALESCE(SUM(potmasculino),0) AS hombre,
                               COALESCE(SUM(potfemenino),0)  AS mujeres,
                               COALESCE(SUM(pottotal),0)     AS total_potencial,
                               COALESCE(SUM(nummesas),0)     AS mesas
                        FROM divipol_presidencial_2026
                        WHERE clase='P' AND coddepto=%s AND codmipio=%s
                        GROUP BY codzona
                        ORDER BY codzona
                    ''', (coddepto, codmipio))
                elif coddepto is not None:
                    cur.execute('''
                        SELECT m.codmipio, m.nommipio,
                               (SELECT COALESCE(SUM(potmasculino),0) FROM divipol_presidencial_2026
                                WHERE clase='P' AND coddepto=m.coddepto AND codmipio=m.codmipio) AS hombre,
                               (SELECT COALESCE(SUM(potfemenino),0) FROM divipol_presidencial_2026
                                WHERE clase='P' AND coddepto=m.coddepto AND codmipio=m.codmipio) AS mujeres,
                               (SELECT COALESCE(SUM(pottotal),0) FROM divipol_presidencial_2026
                                WHERE clase='P' AND coddepto=m.coddepto AND codmipio=m.codmipio) AS total_potencial,
                               (SELECT COALESCE(SUM(nummesas),0) FROM divipol_presidencial_2026
                                WHERE clase='P' AND coddepto=m.coddepto AND codmipio=m.codmipio) AS mesas
                        FROM divipol_presidencial_2026 m
                        WHERE m.clase='M' AND m.coddepto=%s
                        ORDER BY m.codmipio
                    ''', (coddepto,))
                else:
                    cur.execute('''
                        SELECT d.coddepto, d.nomdepto,
                               (SELECT COALESCE(SUM(potmasculino),0) FROM divipol_presidencial_2026
                                WHERE clase='P' AND coddepto=d.coddepto) AS hombre,
                               (SELECT COALESCE(SUM(potfemenino),0) FROM divipol_presidencial_2026
                                WHERE clase='P' AND coddepto=d.coddepto) AS mujeres,
                               (SELECT COALESCE(SUM(pottotal),0) FROM divipol_presidencial_2026
                                WHERE clase='P' AND coddepto=d.coddepto) AS total_potencial,
                               (SELECT COALESCE(SUM(nummesas),0) FROM divipol_presidencial_2026
                                WHERE clase='P' AND coddepto=d.coddepto) AS mesas
                        FROM divipol_presidencial_2026 d
                        WHERE d.clase='D'
                        ORDER BY d.coddepto
                    ''')
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        logger.exception('[divipol]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/divipol/resumen', methods=['GET'])
def divipol_resumen():
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('''
                    SELECT
                      COUNT(*) FILTER (WHERE clase='D') AS deptos,
                      COUNT(*) FILTER (WHERE clase='M') AS municipios,
                      COUNT(*) FILTER (WHERE clase='Z') AS zonas,
                      COUNT(*) FILTER (WHERE clase='P') AS puestos,
                      COALESCE(SUM(nummesas) FILTER (WHERE clase='P'), 0) AS mesas,
                      COALESCE(SUM(pottotal) FILTER (WHERE clase='P'), 0) AS potencial
                    FROM divipol_presidencial_2026
                ''')
                row = cur.fetchone()
        return jsonify({'success': True, 'data': row})
    except Exception as e:
        logger.exception('[divipol/resumen]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/divipol/departamentos', methods=['GET'])
def divipol_departamentos():
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                # Devolver depto con totales agregados
                cur.execute('''
                    SELECT d.coddepto, d.nomdepto,
                           COALESCE((SELECT COUNT(*) FROM divipol_presidencial_2026
                                     WHERE clase='M' AND coddepto = d.coddepto), 0) AS municipios,
                           COALESCE((SELECT SUM(nummesas) FROM divipol_presidencial_2026
                                     WHERE clase='P' AND coddepto = d.coddepto), 0) AS mesas,
                           COALESCE((SELECT SUM(pottotal) FROM divipol_presidencial_2026
                                     WHERE clase='P' AND coddepto = d.coddepto), 0) AS potencial
                    FROM divipol_presidencial_2026 d
                    WHERE d.clase='D'
                    ORDER BY d.coddepto
                ''')
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        logger.exception('[divipol/departamentos]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/divipol/municipios', methods=['GET'])
def divipol_municipios():
    err = _require_session()
    if err: return err
    coddepto = request.args.get('coddepto', type=int)
    if coddepto is None:
        return jsonify({'success': False, 'error': 'coddepto requerido'}), 400
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('''
                    SELECT m.coddepto, m.codmipio, m.nommipio,
                           COALESCE((SELECT COUNT(*) FROM divipol_presidencial_2026
                                     WHERE clase='Z' AND coddepto=m.coddepto AND codmipio=m.codmipio), 0) AS zonas,
                           COALESCE((SELECT COUNT(*) FROM divipol_presidencial_2026
                                     WHERE clase='P' AND coddepto=m.coddepto AND codmipio=m.codmipio), 0) AS puestos,
                           COALESCE((SELECT SUM(nummesas) FROM divipol_presidencial_2026
                                     WHERE clase='P' AND coddepto=m.coddepto AND codmipio=m.codmipio), 0) AS mesas,
                           COALESCE((SELECT SUM(pottotal) FROM divipol_presidencial_2026
                                     WHERE clase='P' AND coddepto=m.coddepto AND codmipio=m.codmipio), 0) AS potencial
                    FROM divipol_presidencial_2026 m
                    WHERE m.clase='M' AND m.coddepto=%s
                    ORDER BY m.codmipio
                ''', (coddepto,))
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        logger.exception('[divipol/municipios]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/divipol/zonas', methods=['GET'])
def divipol_zonas():
    err = _require_session()
    if err: return err
    coddepto = request.args.get('coddepto', type=int)
    codmipio = request.args.get('codmipio', type=int)
    if coddepto is None or codmipio is None:
        return jsonify({'success': False, 'error': 'coddepto y codmipio requeridos'}), 400
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('''
                    SELECT z.coddepto, z.codmipio, z.codzona,
                           COALESCE((SELECT COUNT(*) FROM divipol_presidencial_2026
                                     WHERE clase='P' AND coddepto=z.coddepto AND codmipio=z.codmipio AND codzona=z.codzona), 0) AS puestos,
                           COALESCE((SELECT SUM(nummesas) FROM divipol_presidencial_2026
                                     WHERE clase='P' AND coddepto=z.coddepto AND codmipio=z.codmipio AND codzona=z.codzona), 0) AS mesas,
                           COALESCE((SELECT SUM(pottotal) FROM divipol_presidencial_2026
                                     WHERE clase='P' AND coddepto=z.coddepto AND codmipio=z.codmipio AND codzona=z.codzona), 0) AS potencial
                    FROM divipol_presidencial_2026 z
                    WHERE z.clase='Z' AND z.coddepto=%s AND z.codmipio=%s
                    ORDER BY z.codzona
                ''', (coddepto, codmipio))
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        logger.exception('[divipol/zonas]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/divipol/puestos', methods=['GET'])
def divipol_puestos():
    err = _require_session()
    if err: return err
    coddepto = request.args.get('coddepto', type=int)
    codmipio = request.args.get('codmipio', type=int)
    codzona  = request.args.get('codzona', type=int)
    if coddepto is None or codmipio is None or codzona is None:
        return jsonify({'success': False, 'error': 'coddepto, codmipio, codzona requeridos'}), 400
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('''
                    SELECT iddivipol, coddepto, codmipio, codzona, codpuesto,
                           nompuesto, nummesas, potfemenino, potmasculino, pottotal,
                           jal, nomjal
                    FROM divipol_presidencial_2026
                    WHERE clase='P' AND coddepto=%s AND codmipio=%s AND codzona=%s
                    ORDER BY codpuesto
                ''', (coddepto, codmipio, codzona))
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        logger.exception('[divipol/puestos]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/divipol/mesas', methods=['GET'])
def divipol_mesas():
    err = _require_session()
    if err: return err
    iddivipol = request.args.get('iddivipol', type=int)
    if iddivipol is None:
        return jsonify({'success': False, 'error': 'iddivipol requerido'}), 400
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('''
                    SELECT idmesa, iddivipol, mesa, jornada
                    FROM divipolmesa_presidencial_2026
                    WHERE iddivipol=%s
                    ORDER BY mesa
                ''', (iddivipol,))
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        logger.exception('[divipol/mesas]')
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== MMV PRECONTEO ====================
from datetime import date

@app.route('/api/mmv-preconteo/historial', methods=['GET'])
def mmv_preconteo_historial():
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('''
                    SELECT nombrearchivo, fecha, registros, estado, usuario
                    FROM control_mmv_presidencial_2026
                    ORDER BY fecha DESC, nombrearchivo
                ''')
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        logger.exception('[mmv/historial]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mmv-preconteo/verificar', methods=['POST'])
def mmv_preconteo_verificar():
    err = _require_session()
    if err: return err
    try:
        data = request.get_json() or {}
        nombre = data.get('nombrearchivo', '')
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('''
                    SELECT nombrearchivo, fecha, registros, estado
                    FROM control_mmv_presidencial_2026
                    WHERE nombrearchivo = %s
                ''', (nombre,))
                row = cur.fetchone()
        if row:
            return jsonify({'success': True, 'existe': True, 'data': row})
        return jsonify({'success': True, 'existe': False})
    except Exception as e:
        logger.exception('[mmv/verificar]')
        return jsonify({'success': False, 'error': str(e)}), 500

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
PROCESADO_FOLDER = os.path.join(UPLOAD_FOLDER, 'PROCESADO')
os.makedirs(PROCESADO_FOLDER, exist_ok=True)

# Ruta del repositorio de PDFs E24 Presidencial (formulario por comisión escrutadora)
E24_PRES_BASE_PATH = os.environ.get('E24_PRES_BASE_PATH', '/opt/softwareEscrutinios/E24_PRES')
# Ruta del repositorio de PDFs E14 Presidencial (formulario por mesa)
E14_PRES_BASE_PATH = os.environ.get('E14_PRES_BASE_PATH', '/mnt/elecciones-2026/presidencial')

import re as _re
def _extraer_depto(nombre_archivo, primera_linea=''):
    """Extrae código de depto del nombre (patrón _<dd>_) o de la primera línea (primeros 2 chars)."""
    m = _re.search(r'_(\d{2})_', nombre_archivo)
    if m:
        return m.group(1)
    if primera_linea and len(primera_linea) >= 2 and primera_linea[:2].isdigit():
        return primera_linea[:2]
    return '00'  # depto desconocido

@app.route('/api/mmv-preconteo/cargar', methods=['POST'])
def mmv_preconteo_cargar():
    err = _require_session()
    if err: return err
    try:
        if 'archivo' not in request.files:
            return jsonify({'success': False, 'error': 'No se envió ningún archivo'}), 400
        archivo = request.files['archivo']
        if not archivo.filename:
            return jsonify({'success': False, 'error': 'Nombre de archivo vacío'}), 400

        nombre = archivo.filename
        tiene_encab = request.form.get('tiene_encabezado', 'true') == 'true'

        # Validar duplicado ANTES de procesar
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    'SELECT fecha, registros FROM control_mmv_presidencial_2026 WHERE nombrearchivo = %s',
                    (nombre,))
                ya = cur.fetchone()
        if ya:
            return jsonify({
                'success': False,
                'error': f'El archivo "{nombre}" ya fue cargado el {ya["fecha"]} con {ya["registros"]:,} registros. No se permite cargarlo nuevamente.',
                'duplicado': True,
                'data': {'fecha': str(ya['fecha']), 'registros': ya['registros']}
            }), 409

        contenido_bytes = archivo.read()
        contenido = contenido_bytes.decode('utf-8', errors='replace')
        lineas = [l for l in contenido.replace('\r', '').split('\n') if l.strip()]
        if tiene_encab and lineas:
            lineas = lineas[1:]
        registros = len(lineas)
        usuario = session.get('cedula', '')

        # Determinar depto y guardar copia en PROCESADO/<depto>/
        primera = lineas[0] if lineas else ''
        depto = _extraer_depto(nombre, primera)
        depto_dir = os.path.join(PROCESADO_FOLDER, depto)
        os.makedirs(depto_dir, exist_ok=True)
        ruta_dest = os.path.join(depto_dir, nombre)
        with open(ruta_dest, 'wb') as fout:
            fout.write(contenido_bytes)

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('''
                    INSERT INTO control_mmv_presidencial_2026 (nombrearchivo, fecha, registros, estado, usuario)
                    VALUES (%s, %s, %s, 0, %s)
                ''', (nombre, date.today(), registros, usuario))
                if lineas:
                    cur.executemany(
                        'INSERT INTO preconteo_cargue_presidencial_2026 (dato, archivo) VALUES (%s, %s)',
                        [(l.strip(), nombre) for l in lineas]
                    )
                conn.commit()

        return jsonify({
            'success': True,
            'message': f'Archivo "{nombre}" cargado: {registros:,} registros. Copia guardada en PROCESADO/{depto}/.',
            'data': {'nombrearchivo': nombre, 'fecha': date.today().isoformat(),
                     'registros': registros, 'estado': 0,
                     'depto': depto, 'ruta_procesado': f'PROCESADO/{depto}/{nombre}'}
        })
    except Exception as e:
        logger.exception('[mmv/cargar]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mmv-preconteo/eliminar-batch', methods=['POST'])
def mmv_preconteo_eliminar_batch():
    """Elimina múltiples archivos en una sola operación."""
    err = _require_session()
    if err: return err
    try:
        data = request.get_json() or {}
        archivos = data.get('archivos') or []
        if not archivos or not isinstance(archivos, list):
            return jsonify({'success': False, 'error': 'Lista de archivos requerida'}), 400

        eliminados = 0
        archivos_fisicos_borrados = 0
        errores = []

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('DELETE FROM preconteo_cargue_presidencial_2026 WHERE archivo = ANY(%s)', (archivos,))
                cur.execute('DELETE FROM preconteo_presidencial_2026 WHERE archivo = ANY(%s)', (archivos,))
                cur.execute('DELETE FROM control_mmv_presidencial_2026 WHERE nombrearchivo = ANY(%s)', (archivos,))
                eliminados = cur.rowcount
                conn.commit()

        # Borrar copias físicas
        for nombre in archivos:
            depto = _extraer_depto(nombre)
            ruta = os.path.join(PROCESADO_FOLDER, depto, nombre)
            if os.path.exists(ruta):
                try:
                    os.remove(ruta)
                    archivos_fisicos_borrados += 1
                except OSError as ex:
                    errores.append(f'{nombre}: {ex}')

        return jsonify({
            'success': True,
            'eliminados_bd': eliminados,
            'archivos_fisicos_borrados': archivos_fisicos_borrados,
            'errores': errores,
            'message': f'{eliminados} archivo(s) eliminados de la BD, {archivos_fisicos_borrados} copia(s) físicas borradas.'
        })
    except Exception as e:
        logger.exception('[mmv/eliminar-batch]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mmv-preconteo/<path:nombrearchivo>', methods=['DELETE'])
def mmv_preconteo_eliminar(nombrearchivo):
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('DELETE FROM preconteo_cargue_presidencial_2026 WHERE archivo = %s', (nombrearchivo,))
                cur.execute('DELETE FROM preconteo_presidencial_2026 WHERE archivo = %s', (nombrearchivo,))
                cur.execute('DELETE FROM control_mmv_presidencial_2026 WHERE nombrearchivo = %s', (nombrearchivo,))
                conn.commit()
        # Eliminar copia física en PROCESADO/<depto>/
        depto = _extraer_depto(nombrearchivo)
        ruta_dest = os.path.join(PROCESADO_FOLDER, depto, nombrearchivo)
        if os.path.exists(ruta_dest):
            try: os.remove(ruta_dest)
            except OSError: pass
        return jsonify({'success': True, 'message': f'Archivo "{nombrearchivo}" eliminado.'})
    except Exception as e:
        logger.exception('[mmv/eliminar]')
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== PROCESAR MMV (parser fixed-width) ====================
def _parse_mmv_line(linea):
    """
    Parsea una línea fixed-width MMV Preconteo Presidencial (38 chars).
      coddepto(2) codmipio(3) zona(2) puesto(2) mesa(6) codjal(2)
      comunicado(4) codcircunscripcion(1) codpartido(5) codcandidato(3) votos(8)
    """
    if len(linea) < 38:
        return None
    try:
        return {
            'coddepto':           int(linea[0:2]),
            'codmipio':           int(linea[2:5]),
            'codzona':            int(linea[5:7]),
            'codpuesto':          linea[7:9],
            'mesa':               int(linea[9:15]),
            'codjal':             int(linea[15:17] or 0),
            'comunicado':         int(linea[17:21] or 0),
            'codcircunscripcion': int(linea[21:22] or 0),
            'codpartido':         int(linea[22:27] or 0),
            'codcandidato':       int(linea[27:30] or 0),
            'votos':              int(linea[30:38] or 0),
        }
    except (ValueError, IndexError):
        return None

@app.route('/api/mmv-preconteo/procesar/<path:nombrearchivo>', methods=['POST'])
def mmv_preconteo_procesar(nombrearchivo):
    """Parsea líneas raw del archivo y las mete estructuradas en preconteo_presidencial_2026."""
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    'SELECT estado FROM control_mmv_presidencial_2026 WHERE nombrearchivo = %s',
                    (nombrearchivo,))
                row = cur.fetchone()
                if not row:
                    return jsonify({'success': False, 'error': 'Archivo no encontrado en historial'}), 404
                if row['estado'] == 1:
                    return jsonify({'success': False, 'error': 'El archivo ya fue procesado.'}), 409

                # Limpiar eventuales restos
                cur.execute('DELETE FROM preconteo_presidencial_2026 WHERE archivo = %s', (nombrearchivo,))

                # Leer todas las líneas raw
                cur.execute(
                    'SELECT dato FROM preconteo_cargue_presidencial_2026 WHERE archivo = %s ORDER BY id',
                    (nombrearchivo,))
                lineas = [r['dato'] for r in cur.fetchall()]

                # Parsear
                rows_validos = []
                rows_invalidos = 0
                for ln in lineas:
                    p = _parse_mmv_line(ln)
                    if p is None:
                        rows_invalidos += 1
                        continue
                    rows_validos.append((
                        p['coddepto'], p['codmipio'], p['codzona'], p['codpuesto'],
                        p['mesa'], p['comunicado'], p['codpartido'], p['codcandidato'],
                        p['votos'], nombrearchivo
                    ))

                # Bulk insert
                if rows_validos:
                    with cur.copy(
                        "COPY preconteo_presidencial_2026 (coddepto, codmipio, codzona, codpuesto, mesa, boletin, codpartido, codcandidato, votos, archivo) FROM STDIN"
                    ) as copy:
                        for r in rows_validos:
                            copy.write_row(r)

                # Marcar como procesado
                cur.execute(
                    'UPDATE control_mmv_presidencial_2026 SET estado = 1 WHERE nombrearchivo = %s',
                    (nombrearchivo,))
                conn.commit()

        return jsonify({
            'success': True,
            'message': f'Archivo "{nombrearchivo}" procesado.',
            'data': {
                'lineas_raw': len(lineas),
                'lineas_validas': len(rows_validos),
                'lineas_invalidas': rows_invalidos,
            }
        })
    except Exception as e:
        logger.exception('[mmv/procesar]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/mmv-preconteo/procesar-todos', methods=['POST'])
def mmv_preconteo_procesar_todos():
    """Procesa todos los archivos con estado = 0."""
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT nombrearchivo FROM control_mmv_presidencial_2026 WHERE estado = 0 ORDER BY fecha, nombrearchivo")
                pendientes = [r['nombrearchivo'] for r in cur.fetchall()]
        procesados = 0
        errores = []
        for nombre in pendientes:
            try:
                # Reusa la lógica del endpoint individual
                with app.test_request_context():
                    pass
                with get_db_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute('DELETE FROM preconteo_presidencial_2026 WHERE archivo = %s', (nombre,))
                        cur.execute(
                            'SELECT dato FROM preconteo_cargue_presidencial_2026 WHERE archivo = %s ORDER BY id',
                            (nombre,))
                        lineas = [r['dato'] for r in cur.fetchall()]
                        rows_validos = []
                        for ln in lineas:
                            p = _parse_mmv_line(ln)
                            if p is None: continue
                            rows_validos.append((
                                p['coddepto'], p['codmipio'], p['codzona'], p['codpuesto'],
                                p['mesa'], p['comunicado'], p['codpartido'], p['codcandidato'],
                                p['votos'], nombre))
                        if rows_validos:
                            with cur.copy(
                                "COPY preconteo_presidencial_2026 (coddepto, codmipio, codzona, codpuesto, mesa, boletin, codpartido, codcandidato, votos, archivo) FROM STDIN"
                            ) as copy:
                                for r in rows_validos:
                                    copy.write_row(r)
                        cur.execute(
                            'UPDATE control_mmv_presidencial_2026 SET estado = 1 WHERE nombrearchivo = %s',
                            (nombre,))
                        conn.commit()
                procesados += 1
            except Exception as ex:
                errores.append(f'{nombre}: {ex}')
        return jsonify({
            'success': True,
            'procesados': procesados,
            'pendientes_antes': len(pendientes),
            'errores': errores,
            'message': f'Procesados {procesados} de {len(pendientes)} archivo(s) pendiente(s).'
        })
    except Exception as e:
        logger.exception('[mmv/procesar-todos]')
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== MMV ESCRUTINIO ====================
import tempfile, csv as _csv_mod, time as _time_mod

def _detectar_separador(ruta_archivo):
    """Devuelve ';' o ',' según primera línea no-encabezado."""
    try:
        with open(ruta_archivo, 'r', encoding='utf-8', errors='replace') as f:
            for _ in range(2):
                line = f.readline()
                if line and (';' in line or ',' in line):
                    return ';' if ';' in line else ','
        return ';'
    except Exception:
        return ';'

def _crear_particion_si_falta(cur, facceso):
    """Crea partición de escrutinio_presidencial_2026 para la facceso dada."""
    nombre_p = 'escrutiniopres_' + facceso.replace('-', '')
    cur.execute("SELECT 1 FROM pg_class WHERE relname = %s", (nombre_p,))
    if not cur.fetchone():
        cur.execute(
            f"CREATE TABLE {nombre_p} PARTITION OF escrutinio_presidencial_2026 FOR VALUES IN ('{facceso}')"
        )
        logger.info(f"[escrutinio] partición creada: {nombre_p}")

@app.route('/api/escrutinio/historial', methods=['GET'])
def escrutinio_historial():
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('''
                    SELECT id, facceso, nombrearchivo, registros, fecha, usuario_cargue,
                           tiempo_carga, estado, tamano_mb
                    FROM control_escrutinio_presidencial_2026
                    ORDER BY facceso DESC, nombrearchivo
                ''')
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        logger.exception('[escrutinio/historial]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/escrutinio/verificar', methods=['POST'])
def escrutinio_verificar():
    err = _require_session()
    if err: return err
    try:
        d = request.get_json() or {}
        nombre = d.get('nombrearchivo', '')
        facceso = d.get('facceso', '')
        if not nombre or not facceso:
            return jsonify({'success': False, 'error': 'nombrearchivo y facceso requeridos'}), 400
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    'SELECT facceso, registros, fecha FROM control_escrutinio_presidencial_2026 WHERE nombrearchivo = %s AND facceso = %s',
                    (nombre, facceso))
                row = cur.fetchone()
        if row:
            return jsonify({'success': True, 'existe': True, 'data': row})
        return jsonify({'success': True, 'existe': False})
    except Exception as e:
        logger.exception('[escrutinio/verificar]')
        return jsonify({'success': False, 'error': str(e)}), 500

def _poblar_seguimiento_dia(facceso):
    """Invoca fn_poblar_escrutinio_presidencial al terminar la carga del día."""
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT fn_poblar_escrutinio_presidencial(%s::date)", (facceso,))
                msg = cur.fetchone()
                conn.commit()
        logger.info(f"[seguimiento] {msg}")
        return msg
    except Exception as ex:
        logger.exception('[seguimiento]')
        return f'Error pobloando seguimiento: {ex}'

def _procesar_csv_a_bd(ruta_csv, nombre_archivo, facceso, tiene_encab):
    """Lee CSV streaming → COPY a tabla. Devuelve dict con stats."""
    t0 = _time_mod.time()
    sep = _detectar_separador(ruta_csv)
    tamano_mb = round(os.path.getsize(ruta_csv) / (1024 * 1024), 2)

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            _crear_particion_si_falta(cur, facceso)

            cols = ('facceso, idregistro, codcorporacion, nomcorporacion, '
                    'codcircunscripcion, nomcircunscripcion, coddepto, nomdepto, '
                    'codmipio, nommipio, codzona, nomzona, codpuesto, nompuesto, '
                    'mesa, codcomuna, nomcomuna, codpartido, nompartido, '
                    'cedulacandidato, nomcandidato, codcandidato, votos, archivo')

            registros = 0
            # Auto-detección de encabezado: si la primera línea no tiene un primer campo numérico, es encabezado
            ENCAB_TOKENS = {'id', 'corporacioncodigo', 'departamentocodigo', 'municipiocodigo', 'mesa', 'totalvotos'}
            with cur.copy(f"COPY escrutinio_presidencial_2026 ({cols}) FROM STDIN") as copy:
                with open(ruta_csv, 'r', encoding='utf-8', errors='replace') as fin:
                    rdr = _csv_mod.reader(fin, delimiter=sep)
                    primera = None
                    try: primera = next(rdr)
                    except StopIteration: pass
                    if primera:
                        # ¿Es encabezado? — si los campos contienen tokens conocidos O el primero no es numérico
                        es_encab = False
                        if tiene_encab:
                            es_encab = True
                        else:
                            joined = (','.join(c.lower() for c in primera[:6]) if primera else '')
                            if any(tok in joined for tok in ENCAB_TOKENS):
                                es_encab = True
                                logger.info("[escrutinio] encabezado detectado automáticamente, saltando primera línea")
                        if not es_encab:
                            # Re-procesar la primera línea como dato
                            rows_a_procesar = [primera]
                            for r in rdr: rows_a_procesar.append(r)
                            iter_rows = rows_a_procesar
                        else:
                            iter_rows = rdr
                    else:
                        iter_rows = rdr
                    for row in iter_rows:
                        if not row or len(row) < 22:
                            continue
                        # Normalizar 22 campos del CSV RNEC
                        def _ni(v):
                            try: return int(v) if v not in ('', None) else None
                            except (ValueError, TypeError): return None
                        def _ns(v): return (v or '').strip().replace('\t', ' ')
                        try:
                            valores = [
                                facceso,
                                _ns(row[0]),            # idregistro
                                _ni(row[1]),            # codcorporacion
                                _ns(row[2]),            # nomcorporacion
                                _ni(row[3]),            # codcircunscripcion
                                _ns(row[4]),            # nomcircunscripcion
                                _ni(row[5]),            # coddepto
                                _ns(row[6]),            # nomdepto
                                _ni(row[7]),            # codmipio
                                _ns(row[8]),            # nommipio
                                _ni(row[9]),            # codzona
                                _ns(row[10]),           # nomzona
                                _ns(row[11]),           # codpuesto
                                _ns(row[12]),           # nompuesto
                                _ni(row[13]),           # mesa
                                _ns(row[14]),           # codcomuna
                                _ns(row[15]),           # nomcomuna
                                _ni(row[16]),           # codpartido
                                _ns(row[17]),           # nompartido
                                _ns(row[18]),           # cedulacandidato
                                _ns(row[19]),           # nomcandidato
                                _ni(row[20]),           # codcandidato
                                _ni(row[21]) or 0,      # votos
                                nombre_archivo,
                            ]
                            copy.write_row(valores)
                            registros += 1
                        except (IndexError, ValueError):
                            continue

            conn.commit()

    tiempo = _time_mod.time() - t0
    return {
        'registros': registros, 'tamano_mb': tamano_mb,
        'tiempo_s': round(tiempo, 1),
        'tiempo_carga': f'{int(tiempo//60)}m {int(tiempo%60)}s' if tiempo >= 60 else f'{tiempo:.1f}s',
    }

@app.route('/api/escrutinio/cargar', methods=['POST'])
def escrutinio_cargar():
    err = _require_session()
    if err: return err
    if 'archivo' not in request.files:
        return jsonify({'success': False, 'error': 'No se envió ningún archivo'}), 400
    archivo = request.files['archivo']
    if not archivo.filename:
        return jsonify({'success': False, 'error': 'Nombre de archivo vacío'}), 400
    facceso = request.form.get('facceso', '').strip()
    if not facceso:
        return jsonify({'success': False, 'error': 'Debe indicar fecha de acceso (facceso)'}), 400
    nombre = archivo.filename
    tiene_encab = request.form.get('tiene_encabezado', 'true') == 'true'

    # Validar duplicado
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    'SELECT registros, fecha FROM control_escrutinio_presidencial_2026 WHERE nombrearchivo = %s AND facceso = %s',
                    (nombre, facceso))
                ya = cur.fetchone()
        if ya:
            return jsonify({
                'success': False, 'duplicado': True,
                'error': f'El archivo "{nombre}" ya fue cargado con facceso {facceso} ({ya["registros"]:,} registros).',
                'data': {'fecha': str(ya['fecha']), 'registros': ya['registros']}
            }), 409
    except Exception as e:
        logger.exception('[escrutinio/cargar verificar]')
        return jsonify({'success': False, 'error': str(e)}), 500

    # Guardar archivo en disco temporal (streaming)
    fd, ruta_temp = tempfile.mkstemp(suffix='.csv', prefix='esc_')
    os.close(fd)
    try:
        archivo.save(ruta_temp)
        stats = _procesar_csv_a_bd(ruta_temp, nombre, facceso, tiene_encab)

        # Guardar copia en PROCESADO y registrar en control
        depto = _extraer_depto(nombre)
        depto_dir = os.path.join(PROCESADO_FOLDER, depto)
        os.makedirs(depto_dir, exist_ok=True)
        import shutil
        shutil.copy2(ruta_temp, os.path.join(depto_dir, nombre))

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('''
                    INSERT INTO control_escrutinio_presidencial_2026
                        (facceso, nombrearchivo, registros, usuario_cargue, tiempo_carga, estado, tamano_mb)
                    VALUES (%s, %s, %s, %s, %s, 1, %s)
                ''', (facceso, nombre, stats['registros'], session.get('cedula', ''),
                      stats['tiempo_carga'], stats['tamano_mb']))
                conn.commit()

        # Disparar pobladora de seguimiento para esta facceso
        seg_msg = _poblar_seguimiento_dia(facceso)

        return jsonify({
            'success': True,
            'message': f'Archivo "{nombre}" cargado: {stats["registros"]:,} registros en {stats["tiempo_carga"]} ({stats["tamano_mb"]} MB).',
            'data': {'nombrearchivo': nombre, 'facceso': facceso,
                     'registros': stats['registros'], 'tamano_mb': stats['tamano_mb'],
                     'tiempo_carga': stats['tiempo_carga'], 'depto': depto,
                     'seguimiento': seg_msg}
        })
    except Exception as e:
        logger.exception('[escrutinio/cargar]')
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if os.path.exists(ruta_temp):
            try: os.remove(ruta_temp)
            except OSError: pass

@app.route('/api/escrutinio/cargar-ruta', methods=['POST'])
def escrutinio_cargar_ruta():
    """Carga desde una ruta local (server) sin upload — para archivos grandes."""
    err = _require_session()
    if err: return err
    try:
        d = request.get_json() or {}
        ruta = d.get('ruta', '').strip()
        facceso = d.get('facceso', '').strip()
        tiene_encab = d.get('tiene_encabezado', True)
        if not ruta or not facceso:
            return jsonify({'success': False, 'error': 'ruta y facceso requeridos'}), 400
        if not os.path.exists(ruta):
            return jsonify({'success': False, 'error': f'Archivo no encontrado en server: {ruta}'}), 404
        nombre = os.path.basename(ruta)

        # Validar duplicado
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    'SELECT registros FROM control_escrutinio_presidencial_2026 WHERE nombrearchivo = %s AND facceso = %s',
                    (nombre, facceso))
                if cur.fetchone():
                    return jsonify({
                        'success': False, 'duplicado': True,
                        'error': f'El archivo "{nombre}" ya fue cargado con facceso {facceso}.'
                    }), 409

        stats = _procesar_csv_a_bd(ruta, nombre, facceso, tiene_encab)

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('''
                    INSERT INTO control_escrutinio_presidencial_2026
                        (facceso, nombrearchivo, registros, usuario_cargue, tiempo_carga, estado, tamano_mb)
                    VALUES (%s, %s, %s, %s, %s, 1, %s)
                ''', (facceso, nombre, stats['registros'], session.get('cedula', ''),
                      stats['tiempo_carga'], stats['tamano_mb']))
                conn.commit()

        return jsonify({
            'success': True,
            'message': f'Archivo "{nombre}" cargado por ruta: {stats["registros"]:,} registros en {stats["tiempo_carga"]}.',
            'data': stats
        })
    except Exception as e:
        logger.exception('[escrutinio/cargar-ruta]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/escrutinio/eliminar-batch', methods=['POST'])
def escrutinio_eliminar_batch():
    err = _require_session()
    if err: return err
    try:
        d = request.get_json() or {}
        items = d.get('items') or []  # [{nombrearchivo, facceso}]
        if not items:
            return jsonify({'success': False, 'error': 'items requerido'}), 400
        elim = 0
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                for it in items:
                    cur.execute(
                        'DELETE FROM escrutinio_presidencial_2026 WHERE archivo = %s AND facceso = %s',
                        (it['nombrearchivo'], it['facceso']))
                    cur.execute(
                        'DELETE FROM control_escrutinio_presidencial_2026 WHERE nombrearchivo = %s AND facceso = %s',
                        (it['nombrearchivo'], it['facceso']))
                    elim += cur.rowcount
                conn.commit()
        return jsonify({'success': True, 'eliminados': elim,
                        'message': f'{elim} archivo(s) eliminados.'})
    except Exception as e:
        logger.exception('[escrutinio/eliminar-batch]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/escrutinio/estado-por-dia', methods=['GET'])
def escrutinio_estado_dia():
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('''
                    SELECT facceso,
                           COUNT(*) AS archivos,
                           SUM(registros)::BIGINT AS total_registros,
                           SUM(tamano_mb)::NUMERIC(12,2) AS total_mb,
                           MAX(fecha) AS ultima_carga
                    FROM control_escrutinio_presidencial_2026
                    GROUP BY facceso
                    ORDER BY facceso DESC
                ''')
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        logger.exception('[escrutinio/estado-por-dia]')
        return jsonify({'success': False, 'error': str(e)}), 500

# Etiqueta SQL para mostrar nombre de candidato con casos especiales (996/997/998)
def _nomcandidato_sql(alias_origen):
    """Devuelve expresión SQL CASE para etiqueta legible del candidato.
    alias_origen: 'p' (preconteo) o 's' (seguimiento) — la tabla con codcandidato.
    """
    return (f"CASE {alias_origen}.codcandidato "
            f"WHEN 996 THEN 'VOTO EN BLANCO' "
            f"WHEN 997 THEN 'VOTO NULO' "
            f"WHEN 998 THEN 'VOTO NO MARCADO' "
            f"ELSE COALESCE(c.nomcandidato, '—') END")

# ==================== CONSULTA VOTACIÓN PRECONTEO ====================
@app.route('/api/consulta-preconteo/filtros/departamentos', methods=['GET'])
def consulta_pre_deptos():
    err = _require_session()
    if err: return err
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT coddepto, nomdepto FROM divipol_presidencial_2026
                WHERE clase='D' ORDER BY coddepto
            """)
            return jsonify({'success': True, 'data': cur.fetchall()})

@app.route('/api/consulta-preconteo/filtros/municipios', methods=['GET'])
def consulta_pre_municipios():
    err = _require_session()
    if err: return err
    coddepto = request.args.get('coddepto', type=int)
    if coddepto is None: return jsonify({'success': True, 'data': []})
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT codmipio, nommipio FROM divipol_presidencial_2026
                WHERE clase='M' AND coddepto=%s ORDER BY codmipio
            """, (coddepto,))
            return jsonify({'success': True, 'data': cur.fetchall()})

@app.route('/api/consulta-preconteo/filtros/zonas', methods=['GET'])
def consulta_pre_zonas():
    err = _require_session()
    if err: return err
    coddepto = request.args.get('coddepto', type=int)
    codmipio = request.args.get('codmipio', type=int)
    if coddepto is None or codmipio is None:
        return jsonify({'success': True, 'data': []})
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT codzona FROM divipol_presidencial_2026
                WHERE clase='Z' AND coddepto=%s AND codmipio=%s ORDER BY codzona
            """, (coddepto, codmipio))
            return jsonify({'success': True, 'data': cur.fetchall()})

@app.route('/api/consulta-preconteo/filtros/puestos', methods=['GET'])
def consulta_pre_puestos():
    err = _require_session()
    if err: return err
    coddepto = request.args.get('coddepto', type=int)
    codmipio = request.args.get('codmipio', type=int)
    codzona  = request.args.get('codzona', type=int)
    if coddepto is None or codmipio is None:
        return jsonify({'success': True, 'data': []})
    sql = """SELECT codpuesto, nompuesto FROM divipol_presidencial_2026
             WHERE clase='P' AND coddepto=%s AND codmipio=%s"""
    params = [coddepto, codmipio]
    if codzona is not None:
        sql += ' AND codzona=%s'; params.append(codzona)
    sql += ' ORDER BY codpuesto'
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            return jsonify({'success': True, 'data': cur.fetchall()})

@app.route('/api/consulta-preconteo/filtros/candidatos', methods=['GET'])
def consulta_pre_candidatos():
    err = _require_session()
    if err: return err
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT c.codcandidato, c.nomcandidato, c.codpartido, p.nompartido
                FROM candidatos_presidencial_2026 c
                LEFT JOIN partidos_presidencial_2026 p USING (codpartido)
                ORDER BY c.num_tarjeton, c.codcandidato
            """)
            return jsonify({'success': True, 'data': cur.fetchall()})

@app.route('/api/consulta-preconteo/consultar', methods=['POST'])
def consulta_pre_consultar():
    """Agregación de votos según filtros. Adapta el tipo de respuesta por nivel."""
    err = _require_session()
    if err: return err
    try:
        data = request.get_json() or {}
        cd  = data.get('coddepto')
        cm  = data.get('codmipio')
        cz  = data.get('codzona')
        cp  = data.get('codpuesto')
        ccd = data.get('codcandidato')
        nomcand_sql = _nomcandidato_sql('p')

        where, params = [], []
        if cd  is not None: where.append('p.coddepto = %s');     params.append(int(cd))
        if cm  is not None: where.append('p.codmipio = %s');     params.append(int(cm))
        if cz  is not None: where.append('p.codzona = %s');      params.append(int(cz))
        if cp  is not None: where.append('p.codpuesto = %s');    params.append(str(cp).zfill(2))
        if ccd is not None: where.append('p.codcandidato = %s'); params.append(int(ccd))
        where.append('p.votos > 0')
        where_sql = ' AND '.join(where) if where else '1=1'

        # Detectar nivel para decidir agrupación
        sin_filtros = cd is None and cm is None and cz is None and cp is None and ccd is None
        solo_cand   = ccd is not None and cd is None and cm is None and cz is None and cp is None
        depto_only  = cd is not None and cm is None and cz is None and cp is None
        mpio_only   = cd is not None and cm is not None and cz is None and cp is None
        zona_only   = cd is not None and cm is not None and cz is not None and cp is None
        puesto_only = cd is not None and cm is not None and cz is not None and cp is not None

        base_join = """preconteo_presidencial_2026 p
            LEFT JOIN candidatos_presidencial_2026 c
                   ON c.codcandidato = p.codcandidato AND c.codpartido = p.codpartido
            LEFT JOIN partidos_presidencial_2026 pt
                   ON pt.codpartido = p.codpartido"""

        # Adaptar el GROUP BY al nivel
        if sin_filtros or solo_cand:
            sql = f"""SELECT p.codcandidato,
                            {nomcand_sql} AS nomcandidato,
                            COALESCE(pt.nompartido, '—') AS nompartido,
                            SUM(p.votos) AS total_votos
                     FROM {base_join}
                     WHERE {where_sql}
                     GROUP BY p.codcandidato, c.nomcandidato, pt.nompartido
                     ORDER BY total_votos DESC LIMIT 500"""
            tipo = 'resumen_nacional' if sin_filtros else 'candidato_nacional'

        elif depto_only:
            sql = f"""SELECT p.codcandidato,
                            {nomcand_sql} AS nomcandidato,
                            COALESCE(pt.nompartido, '—') AS nompartido,
                            SUM(p.votos) AS total_votos
                     FROM {base_join}
                     WHERE {where_sql}
                     GROUP BY p.codcandidato, c.nomcandidato, pt.nompartido
                     ORDER BY total_votos DESC LIMIT 500"""
            tipo = 'por_departamento'

        elif mpio_only:
            sql = f"""SELECT p.codcandidato,
                            {nomcand_sql} AS nomcandidato,
                            COALESCE(pt.nompartido, '—') AS nompartido,
                            SUM(p.votos) AS total_votos
                     FROM {base_join}
                     WHERE {where_sql}
                     GROUP BY p.codcandidato, c.nomcandidato, pt.nompartido
                     ORDER BY total_votos DESC LIMIT 500"""
            tipo = 'por_municipio'

        elif zona_only:
            sql = f"""SELECT p.codpuesto,
                            (SELECT nompuesto FROM divipol_presidencial_2026 dv
                             WHERE dv.clase='P' AND dv.coddepto=p.coddepto
                               AND dv.codmipio=p.codmipio AND dv.codzona=p.codzona
                               AND dv.codpuesto=p.codpuesto LIMIT 1) AS nompuesto,
                            p.codcandidato,
                            {nomcand_sql} AS nomcandidato,
                            SUM(p.votos) AS total_votos
                     FROM {base_join}
                     WHERE {where_sql}
                     GROUP BY p.codpuesto, p.codcandidato, c.nomcandidato, p.coddepto, p.codmipio, p.codzona
                     ORDER BY p.codpuesto, total_votos DESC LIMIT 1000"""
            tipo = 'por_zona'

        elif puesto_only:
            sql = f"""SELECT p.mesa, p.codcandidato,
                            {nomcand_sql} AS nomcandidato,
                            COALESCE(pt.nompartido, '—') AS nompartido,
                            SUM(p.votos) AS total_votos
                     FROM {base_join}
                     WHERE {where_sql}
                     GROUP BY p.mesa, p.codcandidato, c.nomcandidato, pt.nompartido
                     ORDER BY p.mesa, total_votos DESC LIMIT 5000"""
            tipo = 'por_puesto_mesas'

        else:
            sql = f"""SELECT p.coddepto, p.codmipio, p.codzona, p.codpuesto, p.mesa,
                            p.codcandidato,
                            {nomcand_sql} AS nomcandidato,
                            SUM(p.votos) AS total_votos
                     FROM {base_join}
                     WHERE {where_sql}
                     GROUP BY p.coddepto, p.codmipio, p.codzona, p.codpuesto, p.mesa,
                              p.codcandidato, c.nomcandidato
                     ORDER BY total_votos DESC LIMIT 500"""
            tipo = 'detalle'

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(sql, params)
                rows = cur.fetchall()

        return jsonify({'success': True, 'tipo_resultado': tipo, 'data': rows, 'total': len(rows)})
    except Exception as e:
        logger.exception('[consulta-preconteo/consultar]')
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== INVESTIGACIONES PRESIDENCIAL ====================

@app.route('/api/investigaciones-pres/comparar', methods=['POST'])
def inv_pres_comparar():
    """Compara 2 candidatos mesa a mesa: preconteo vs último día escrutinio."""
    err = _require_session()
    if err: return err
    try:
        d = request.get_json() or {}
        codcand1 = int(d.get('codcandidato1') or 0)
        codcand2 = int(d.get('codcandidato2') or 0)
        cd = d.get('coddepto')
        cm = d.get('codmipio')
        numdia = d.get('numdia')
        solo_pierde1 = bool(d.get('solo_pierde1', False))
        solo_gana1   = bool(d.get('solo_gana1', False))
        solo_pierde2 = bool(d.get('solo_pierde2', False))
        solo_gana2   = bool(d.get('solo_gana2', False))
        filtro_recl  = d.get('filtro_reclamacion', 'todas')   # todas/con_evidencia/sin_evidencia
        pagina = max(1, int(d.get('pagina', 1)))
        por_pagina = min(2000, max(20, int(d.get('por_pagina', 100))))

        if not codcand1 or not codcand2:
            return jsonify({'success': False, 'error': 'Debe seleccionar ambos candidatos'}), 400

        ultimo = _ultimo_dia_seguimiento()
        if ultimo == 0:
            return jsonify({'success': True, 'filas': [], 'total': 0,
                            'mensaje': 'No hay días de escrutinio procesados.'})
        numdia_use = int(numdia) if numdia else ultimo
        expr = _build_ultimo_valor_expr(numdia_use)

        # Obtener nombres y partidos de los candidatos
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT codcandidato, codpartido, nomcandidato,
                           (SELECT nompartido FROM partidos_presidencial_2026 WHERE codpartido=c.codpartido) AS nompartido
                    FROM candidatos_presidencial_2026 c
                    WHERE codcandidato = ANY(%s)
                """, ([codcand1, codcand2],))
                rows = cur.fetchall()
        cand_info = {r['codcandidato']: r for r in rows}
        nom1 = (cand_info.get(codcand1) or {}).get('nomcandidato') or f'Cand {codcand1}'
        nom2 = (cand_info.get(codcand2) or {}).get('nomcandidato') or f'Cand {codcand2}'
        part1 = (cand_info.get(codcand1) or {}).get('nompartido') or '—'
        part2 = (cand_info.get(codcand2) or {}).get('nompartido') or '—'

        # Where geo
        where_geo, params_geo = [], []
        if cd is not None: where_geo.append('dv.coddepto=%s');  params_geo.append(int(cd))
        if cm is not None: where_geo.append('dv.codmipio=%s');  params_geo.append(int(cm))
        geo_sql = ' AND '.join(where_geo) if where_geo else '1=1'

        # Sub-CTE: votos por candidato por mesa (preconteo + dia)
        sql = f"""
            WITH lado1 AS (
                SELECT s.idmesa, s.preconteo, {expr} AS dia_valor
                FROM seguimiento_escrutinio_presidencial_2026 s
                WHERE s.codcandidato = %s
            ),
            lado2 AS (
                SELECT s.idmesa, s.preconteo, {expr} AS dia_valor
                FROM seguimiento_escrutinio_presidencial_2026 s
                WHERE s.codcandidato = %s
            ),
            mesas AS (
                SELECT l1.idmesa,
                       l1.preconteo AS preconteo1, l1.dia_valor AS dia_valor1,
                       (l1.dia_valor - l1.preconteo) AS diferencia1,
                       l2.preconteo AS preconteo2, l2.dia_valor AS dia_valor2,
                       (l2.dia_valor - l2.preconteo) AS diferencia2,
                       dv.coddepto, dv.nomdepto, dv.codmipio, dv.nommipio,
                       dv.codzona, dv.codpuesto, dv.nompuesto, dm.mesa
                FROM lado1 l1
                JOIN lado2 l2 ON l2.idmesa = l1.idmesa
                JOIN divipolmesa_presidencial_2026 dm ON dm.idmesa = l1.idmesa
                JOIN divipol_presidencial_2026 dv ON dv.iddivipol = dm.iddivipol AND dv.clase='P'
                WHERE {geo_sql}
            )
            SELECT *,
                (SELECT COUNT(*) FROM evidencias_presidencial_2026 e
                  WHERE e.idmesa=mesas.idmesa AND e.codcandidato=%s
                    AND COALESCE(e.tipo_formulario,'') NOT IN ('NO_E14','SIN_EVIDENCIA')) AS num_evidencias1,
                (SELECT usuario FROM evidencias_presidencial_2026 e
                  WHERE e.idmesa=mesas.idmesa AND e.codcandidato=%s
                    AND COALESCE(e.tipo_formulario,'') NOT IN ('NO_E14','SIN_EVIDENCIA')
                  ORDER BY fecha ASC LIMIT 1) AS usuario_ev1,
                (SELECT COUNT(*) FROM evidencias_presidencial_2026 e
                  WHERE e.idmesa=mesas.idmesa AND e.codcandidato=%s
                    AND COALESCE(e.tipo_formulario,'') NOT IN ('NO_E14','SIN_EVIDENCIA')) AS num_evidencias2,
                (SELECT usuario FROM evidencias_presidencial_2026 e
                  WHERE e.idmesa=mesas.idmesa AND e.codcandidato=%s
                    AND COALESCE(e.tipo_formulario,'') NOT IN ('NO_E14','SIN_EVIDENCIA')
                  ORDER BY fecha ASC LIMIT 1) AS usuario_ev2,
                (SELECT usuario FROM reservas_mesa_presidencial_2026 r
                  WHERE r.idmesa=mesas.idmesa AND r.lado=1 LIMIT 1) AS reserva1,
                (SELECT usuario FROM reservas_mesa_presidencial_2026 r
                  WHERE r.idmesa=mesas.idmesa AND r.lado=2 LIMIT 1) AS reserva2,
                (SELECT 1 FROM evidencias_presidencial_2026 e
                  WHERE e.idmesa=mesas.idmesa AND e.tipo_formulario='SIN_EVIDENCIA' LIMIT 1) AS sin_evidencia
            FROM mesas
        """

        # Filtros pierde/gana (OR)
        filtros_pg = []
        if solo_pierde1: filtros_pg.append('diferencia1 < 0')
        if solo_gana1:   filtros_pg.append('diferencia1 > 0')
        if solo_pierde2: filtros_pg.append('diferencia2 < 0')
        if solo_gana2:   filtros_pg.append('diferencia2 > 0')
        having_pg = ''
        if filtros_pg:
            sql = f"SELECT * FROM ({sql}) sub WHERE (" + ' OR '.join(filtros_pg) + ")"

        # Filtro reclamacion
        if filtro_recl == 'con_evidencia':
            sql = f"SELECT * FROM ({sql}) sub2 WHERE num_evidencias1 > 0 OR num_evidencias2 > 0"
        elif filtro_recl == 'sin_evidencia':
            sql = f"SELECT * FROM ({sql}) sub2 WHERE num_evidencias1 = 0 AND num_evidencias2 = 0 AND sin_evidencia IS NULL"

        # Total
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(f"SELECT COUNT(*) AS n FROM ({sql}) tot",
                            [codcand1, codcand2] + params_geo + [codcand1, codcand1, codcand2, codcand2])
                total = cur.fetchone()['n']

                # Página
                offset = (pagina - 1) * por_pagina
                sql_pag = sql + ' ORDER BY ABS(COALESCE(diferencia1,0)) + ABS(COALESCE(diferencia2,0)) DESC LIMIT %s OFFSET %s'
                cur.execute(sql_pag,
                            [codcand1, codcand2] + params_geo +
                            [codcand1, codcand1, codcand2, codcand2, por_pagina, offset])
                filas = cur.fetchall()

        return jsonify({
            'success': True,
            'total': total, 'pagina': pagina, 'por_pagina': por_pagina,
            'paginas': (total + por_pagina - 1) // por_pagina,
            'numdia': numdia_use, 'ultimo_dia': ultimo,
            'candidato1': nom1, 'partido1': part1, 'codcandidato1': codcand1,
            'candidato2': nom2, 'partido2': part2, 'codcandidato2': codcand2,
            'filas': filas
        })
    except Exception as e:
        logger.exception('[inv-pres/comparar]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/investigaciones-pres/mesas-cero', methods=['POST'])
def inv_pres_mesas_cero():
    """Busca mesas donde un candidato tiene 0 votos según la fuente:
      - 'escrutinio' (default): 0 en día N, mesa con votos en escrutinio
      - 'preconteo': 0 en preconteo, mesa con votos en preconteo
      - 'perdidos':  preconteo > 0  PERO  escrutinio día N = 0 (votos que desaparecieron)
    """
    err = _require_session()
    if err: return err
    try:
        d = request.get_json() or {}
        codcand = int(d.get('codcandidato') or 0)
        cd = d.get('coddepto')
        cm = d.get('codmipio')
        numdia = d.get('numdia')
        fuente = (d.get('fuente') or 'escrutinio').lower()
        pagina = max(1, int(d.get('pagina', 1)))
        por_pagina = min(2000, max(20, int(d.get('por_pagina', 100))))

        if not codcand:
            return jsonify({'success': False, 'error': 'codcandidato requerido'}), 400
        if fuente not in ('escrutinio', 'preconteo', 'perdidos'):
            return jsonify({'success': False, 'error': "fuente inválida (use escrutinio/preconteo/perdidos)"}), 400

        ultimo = _ultimo_dia_seguimiento()
        usa_escr = fuente in ('escrutinio', 'perdidos')
        if usa_escr and ultimo == 0:
            return jsonify({'success': True, 'filas': [], 'total': 0,
                            'mensaje': 'No hay días procesados (escrutinio).'})
        numdia_use = int(numdia) if numdia else ultimo
        expr_dia = _build_ultimo_valor_expr(numdia_use) if usa_escr else '0'

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT codcandidato, codpartido, nomcandidato,
                           (SELECT nompartido FROM partidos_presidencial_2026 WHERE codpartido=c.codpartido) AS nompartido
                    FROM candidatos_presidencial_2026 c
                    WHERE codcandidato = %s
                """, (codcand,))
                ci = cur.fetchone()
        if not ci:
            return jsonify({'success': False, 'error': 'Candidato no encontrado'}), 404

        where_geo, params_geo = [], []
        if cd is not None: where_geo.append('dv.coddepto=%s'); params_geo.append(int(cd))
        if cm is not None: where_geo.append('dv.codmipio=%s'); params_geo.append(int(cm))
        geo_sql = ' AND '.join(where_geo) if where_geo else '1=1'

        # CTEs según fuente
        if fuente == 'escrutinio':
            # mesas con votos en escrutinio + candidato con 0 votos en escrutinio
            cte_mesas = f"""SELECT DISTINCT idmesa FROM seguimiento_escrutinio_presidencial_2026 WHERE ({expr_dia}) > 0"""
            cte_cand  = f"""SELECT idmesa, SUM({expr_dia}) AS votos FROM seguimiento_escrutinio_presidencial_2026
                              WHERE codcandidato=%s GROUP BY idmesa"""
            label_votos = 'Votos cand. (escrutinio)'
        elif fuente == 'preconteo':
            cte_mesas = """SELECT DISTINCT idmesa FROM seguimiento_escrutinio_presidencial_2026 WHERE COALESCE(preconteo,0) > 0"""
            cte_cand  = """SELECT idmesa, SUM(COALESCE(preconteo,0)) AS votos FROM seguimiento_escrutinio_presidencial_2026
                            WHERE codcandidato=%s GROUP BY idmesa"""
            label_votos = 'Votos cand. (preconteo)'
        else:  # perdidos
            # candidato tenía votos en preconteo PERO 0 en escrutinio día N
            cte_mesas = f"""SELECT DISTINCT idmesa FROM seguimiento_escrutinio_presidencial_2026 WHERE ({expr_dia}) > 0"""
            cte_cand  = f"""SELECT idmesa,
                                   SUM(COALESCE(preconteo,0)) AS preconteo_cand,
                                   SUM({expr_dia}) AS votos
                            FROM seguimiento_escrutinio_presidencial_2026
                            WHERE codcandidato=%s GROUP BY idmesa"""
            label_votos = 'Preconteo cand → Día N'

        # Construcción de query
        if fuente == 'perdidos':
            sql = f"""
                WITH mesas_con_voto AS ({cte_mesas}),
                     voto_cand     AS ({cte_cand})
                SELECT mv.idmesa,
                       dv.coddepto, dv.nomdepto, dv.codmipio, dv.nommipio,
                       dv.codzona, dv.codpuesto, dv.nompuesto, dm.mesa,
                       COALESCE(vc.preconteo_cand, 0) AS preconteo_cand,
                       COALESCE(vc.votos, 0)          AS votos_candidato,
                       (SELECT COUNT(*) FROM evidencias_presidencial_2026 e
                         WHERE e.idmesa=mv.idmesa AND e.codcandidato=%s
                           AND COALESCE(e.tipo_formulario,'') NOT IN ('NO_E14','SIN_EVIDENCIA')) AS num_evidencias,
                       (SELECT 1 FROM evidencias_presidencial_2026 e
                         WHERE e.idmesa=mv.idmesa AND e.tipo_formulario='SIN_EVIDENCIA' LIMIT 1) AS sin_evidencia,
                       (SELECT usuario FROM reservas_mesa_presidencial_2026 r
                         WHERE r.idmesa=mv.idmesa AND r.codcandidato=%s LIMIT 1) AS reserva
                FROM mesas_con_voto mv
                JOIN divipolmesa_presidencial_2026 dm ON dm.idmesa = mv.idmesa
                JOIN divipol_presidencial_2026 dv ON dv.iddivipol = dm.iddivipol AND dv.clase='P'
                LEFT JOIN voto_cand vc ON vc.idmesa = mv.idmesa
                WHERE COALESCE(vc.preconteo_cand, 0) > 0
                  AND COALESCE(vc.votos, 0) = 0
                  AND {geo_sql}
            """
        else:
            sql = f"""
                WITH mesas_con_voto AS ({cte_mesas}),
                     voto_cand     AS ({cte_cand})
                SELECT mv.idmesa,
                       dv.coddepto, dv.nomdepto, dv.codmipio, dv.nommipio,
                       dv.codzona, dv.codpuesto, dv.nompuesto, dm.mesa,
                       COALESCE(vc.votos, 0) AS votos_candidato,
                       (SELECT COUNT(*) FROM evidencias_presidencial_2026 e
                         WHERE e.idmesa=mv.idmesa AND e.codcandidato=%s
                           AND COALESCE(e.tipo_formulario,'') NOT IN ('NO_E14','SIN_EVIDENCIA')) AS num_evidencias,
                       (SELECT 1 FROM evidencias_presidencial_2026 e
                         WHERE e.idmesa=mv.idmesa AND e.tipo_formulario='SIN_EVIDENCIA' LIMIT 1) AS sin_evidencia,
                       (SELECT usuario FROM reservas_mesa_presidencial_2026 r
                         WHERE r.idmesa=mv.idmesa AND r.codcandidato=%s LIMIT 1) AS reserva
                FROM mesas_con_voto mv
                JOIN divipolmesa_presidencial_2026 dm ON dm.idmesa = mv.idmesa
                JOIN divipol_presidencial_2026 dv ON dv.iddivipol = dm.iddivipol AND dv.clase='P'
                LEFT JOIN voto_cand vc ON vc.idmesa = mv.idmesa
                WHERE COALESCE(vc.votos, 0) = 0
                  AND {geo_sql}
            """

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                params = [codcand, codcand, codcand] + params_geo
                cur.execute(f"SELECT COUNT(*) AS n FROM ({sql}) tot", params)
                total = cur.fetchone()['n']
                offset = (pagina - 1) * por_pagina
                cur.execute(sql + " ORDER BY dv.coddepto, dv.codmipio, dm.mesa LIMIT %s OFFSET %s",
                            params + [por_pagina, offset])
                filas = cur.fetchall()

        return jsonify({
            'success': True,
            'fuente': fuente, 'label_votos': label_votos,
            'total': total, 'pagina': pagina, 'por_pagina': por_pagina,
            'paginas': (total + por_pagina - 1) // por_pagina,
            'numdia': numdia_use, 'ultimo_dia': ultimo,
            'codcandidato': codcand, 'nomcandidato': ci['nomcandidato'],
            'codpartido': ci['codpartido'], 'nompartido': ci['nompartido'],
            'filas': filas
        })
    except Exception as e:
        logger.exception('[inv-pres/mesas-cero]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/investigaciones-pres/crear-mesas-cero', methods=['POST'])
def inv_pres_crear_mesas_cero():
    """Crea investigación de mesas cero (1 candidato, lado B vacío)."""
    err = _require_session()
    if err: return err
    try:
        d = request.get_json() or {}
        mesas = d.get('mesas') or []
        codcand = int(d.get('codcandidato') or 0)
        nom = d.get('nomcandidato', '')
        codpart = d.get('codpartido')
        nompart = d.get('nompartido', '')
        numdia = int(d.get('numdia') or 0)
        usuario = session.get('cedula', '')

        if not mesas or not codcand:
            return jsonify({'success': False, 'error': 'mesas + codcandidato requeridos'}), 400

        creadas = 0
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                for m in mesas:
                    cur.execute("""
                        INSERT INTO investigaciones_presidencial_2026 (
                            idmesa, nomdepto, nommipio, nompuesto, mesa,
                            coddepto, codmipio, codzona, codpuesto,
                            codcandidato1, nom_candidato1, codpartido1, nom_partido1,
                            preconteo1, dia_valor1, diferencia1,
                            codcandidato2, nom_candidato2,
                            numdia, usuario_creacion
                        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,
                                  %s,%s,%s,%s,%s,%s,%s,
                                  %s,%s,%s,%s)
                        ON CONFLICT (idmesa, codcandidato1, codcandidato2) DO NOTHING
                    """, (m.get('idmesa'), m.get('nomdepto'), m.get('nommipio'),
                          m.get('nompuesto'), m.get('mesa'),
                          m.get('coddepto'), m.get('codmipio'), m.get('codzona'), m.get('codpuesto'),
                          codcand, nom, codpart, nompart,
                          0, 0, 0,
                          0, 'MESA CERO',
                          numdia, usuario))
                    creadas += cur.rowcount
                conn.commit()
        return jsonify({'success': True, 'creadas': creadas, 'total': len(mesas)})
    except Exception as e:
        logger.exception('[inv-pres/crear-mesas-cero]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/investigaciones-pres/crear', methods=['POST'])
def inv_pres_crear():
    """Crea registros de investigación a partir de mesas comparadas."""
    err = _require_session()
    if err: return err
    try:
        d = request.get_json() or {}
        mesas = d.get('mesas') or []
        codcand1 = int(d.get('codcandidato1') or 0)
        codcand2 = int(d.get('codcandidato2') or 0)
        nom1 = d.get('candidato1', '')
        nom2 = d.get('candidato2', '')
        part1 = d.get('partido1', '')
        part2 = d.get('partido2', '')
        codpart1 = d.get('codpartido1')
        codpart2 = d.get('codpartido2')
        numdia = int(d.get('numdia') or 0)
        usuario = session.get('cedula', '')

        if not mesas or not codcand1 or not codcand2:
            return jsonify({'success': False, 'error': 'mesas + candidatos requeridos'}), 400

        creadas = 0
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                for m in mesas:
                    cur.execute("""
                        INSERT INTO investigaciones_presidencial_2026 (
                            idmesa, nomdepto, nommipio, nompuesto, mesa,
                            coddepto, codmipio, codzona, codpuesto,
                            codcandidato1, nom_candidato1, codpartido1, nom_partido1,
                            preconteo1, dia_valor1, diferencia1,
                            codcandidato2, nom_candidato2, codpartido2, nom_partido2,
                            preconteo2, dia_valor2, diferencia2,
                            numdia, usuario_creacion
                        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,
                                  %s,%s,%s,%s,%s,%s,%s,
                                  %s,%s,%s,%s,%s,%s,%s,
                                  %s,%s)
                        ON CONFLICT (idmesa, codcandidato1, codcandidato2) DO NOTHING
                    """, (m.get('idmesa'), m.get('nomdepto'), m.get('nommipio'), m.get('nompuesto'), m.get('mesa'),
                          m.get('coddepto'), m.get('codmipio'), m.get('codzona'), m.get('codpuesto'),
                          codcand1, nom1, codpart1, part1,
                          m.get('preconteo1', 0), m.get('dia_valor1', 0), m.get('diferencia1', 0),
                          codcand2, nom2, codpart2, part2,
                          m.get('preconteo2', 0), m.get('dia_valor2', 0), m.get('diferencia2', 0),
                          numdia, usuario))
                    creadas += cur.rowcount
                conn.commit()
        return jsonify({'success': True, 'creadas': creadas, 'total': len(mesas)})
    except Exception as e:
        logger.exception('[inv-pres/crear]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/investigaciones-pres/listar-agrupadas', methods=['GET'])
def inv_pres_listar_agrupadas():
    """Lista investigaciones agrupadas por par (cand1, cand2) y depto."""
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT i.codcandidato1, i.nom_candidato1, i.codpartido1, i.nom_partido1,
                           i.codcandidato2, i.nom_candidato2, i.codpartido2, i.nom_partido2,
                           i.coddepto, MIN(i.nomdepto) AS nomdepto,
                           COUNT(*) AS num_mesas,
                           COUNT(DISTINCT i.codmipio) AS num_mpios,
                           SUM(CASE WHEN EXISTS (
                               SELECT 1 FROM evidencias_presidencial_2026 e
                               WHERE e.idmesa=i.idmesa
                                 AND e.codcandidato IN (i.codcandidato1, i.codcandidato2)
                                 AND COALESCE(e.tipo_formulario,'') NOT IN ('NO_E14','SIN_EVIDENCIA')
                           ) THEN 1 ELSE 0 END) AS mesas_con_evidencia,
                           STRING_AGG(DISTINCT i.nommipio, ', ' ORDER BY i.nommipio) AS municipios
                    FROM investigaciones_presidencial_2026 i
                    GROUP BY i.codcandidato1, i.nom_candidato1, i.codpartido1, i.nom_partido1,
                             i.codcandidato2, i.nom_candidato2, i.codpartido2, i.nom_partido2, i.coddepto
                    ORDER BY i.coddepto, num_mesas DESC
                """)
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows, 'total': len(rows)})
    except Exception as e:
        logger.exception('[inv-pres/listar]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/investigaciones-pres/detalle-grupo', methods=['GET'])
def inv_pres_detalle_grupo():
    """Detalle de mesas de un grupo (par candidatos × depto)."""
    err = _require_session()
    if err: return err
    try:
        coddepto = request.args.get('coddepto', type=int)
        ccd1 = request.args.get('codcandidato1', type=int)
        ccd2 = request.args.get('codcandidato2', type=int)
        pagina = max(1, int(request.args.get('pagina', 1)))
        por_pagina = min(500, max(20, int(request.args.get('por_pagina', 100))))
        if not ccd1 or not ccd2:
            return jsonify({'success': False, 'error': 'codcandidato1 y codcandidato2 requeridos'}), 400

        where = ['codcandidato1=%s', 'codcandidato2=%s']
        params = [ccd1, ccd2]
        if coddepto is not None:
            where.append('coddepto=%s'); params.append(coddepto)

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(f"SELECT COUNT(*) AS n FROM investigaciones_presidencial_2026 WHERE {' AND '.join(where)}", params)
                total = cur.fetchone()['n']
                offset = (pagina - 1) * por_pagina
                cur.execute(f"""
                    SELECT i.idmesa, i.coddepto, i.nomdepto, i.codmipio, i.nommipio,
                           i.codzona, i.codpuesto, i.nompuesto, i.mesa,
                           i.preconteo1, i.dia_valor1, i.diferencia1,
                           i.preconteo2, i.dia_valor2, i.diferencia2,
                           i.estado_reclamacion, i.usuario_asignado,
                           (SELECT COUNT(*) FROM evidencias_presidencial_2026 e
                             WHERE e.idmesa=i.idmesa
                               AND e.codcandidato IN (i.codcandidato1, i.codcandidato2)
                               AND COALESCE(e.tipo_formulario,'') NOT IN ('NO_E14','SIN_EVIDENCIA')) AS num_evidencias
                    FROM investigaciones_presidencial_2026 i
                    WHERE {' AND '.join(where)}
                    ORDER BY i.nommipio, i.codzona, i.codpuesto, i.mesa
                    LIMIT %s OFFSET %s
                """, params + [por_pagina, offset])
                filas = cur.fetchall()
        return jsonify({'success': True, 'data': filas, 'total': total,
                        'pagina': pagina, 'por_pagina': por_pagina,
                        'paginas': (total + por_pagina - 1) // por_pagina})
    except Exception as e:
        logger.exception('[inv-pres/detalle]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/investigaciones-pres/eliminar-grupo', methods=['POST'])
def inv_pres_eliminar_grupo():
    """Elimina todas las mesas de un grupo (par candidatos [× depto])."""
    err = _require_session()
    if err: return err
    try:
        d = request.get_json() or {}
        ccd1 = int(d.get('codcandidato1') or 0)
        ccd2 = int(d.get('codcandidato2') or 0)
        coddepto = d.get('coddepto')
        if not ccd1 or not ccd2:
            return jsonify({'success': False, 'error': 'candidatos requeridos'}), 400
        where = ['codcandidato1=%s', 'codcandidato2=%s']
        params = [ccd1, ccd2]
        if coddepto is not None:
            where.append('coddepto=%s'); params.append(int(coddepto))
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(f"DELETE FROM investigaciones_presidencial_2026 WHERE {' AND '.join(where)}", params)
                eliminadas = cur.rowcount
                conn.commit()
        return jsonify({'success': True, 'eliminadas': eliminadas})
    except Exception as e:
        logger.exception('[inv-pres/eliminar]')
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== VISOR E14 PRESIDENCIAL ====================
import base64 as _b64

def _e14_pres_path_token(ruta_relativa):
    return _b64.urlsafe_b64encode(ruta_relativa.encode()).decode().rstrip('=')

def _e14_pres_decode_token(token):
    padded = token + '=' * (-len(token) % 4)
    return _b64.urlsafe_b64decode(padded.encode()).decode()

@app.route('/api/e14-pres/buscar', methods=['GET'])
def e14_pres_buscar():
    """Busca E14 disponibles para una divipol dada (consulta e14_index_presidencial)."""
    err = _require_session()
    if err: return err
    try:
        cd = request.args.get('coddepto', type=int)
        cm = request.args.get('codmipio', type=int)
        cz = request.args.get('codzona', type=int)
        cp = request.args.get('codpuesto')
        mesa = request.args.get('mesa')
        if not all([cd is not None, cm is not None, cz is not None, cp]):
            return jsonify({'success': False, 'error': 'Faltan parámetros de divipol'}), 400
        params = [str(cd).zfill(2), str(cm).zfill(3), str(cz).zfill(3), str(cp).zfill(2)]
        sql = """SELECT nombre_archivo, corporacion, fuente, ruta_archivo
                 FROM e14_index_presidencial
                 WHERE coddepto=%s AND codmipio=%s AND codzona=%s AND codpuesto=%s"""
        if mesa:
            sql += ' AND mesa = %s'; params.append(str(mesa).zfill(3))
        sql += ' ORDER BY fuente, mesa'
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(sql, params)
                rows = cur.fetchall()
        out = {'claveros': [], 'delegados': [], 'transmision': []}
        for r in rows:
            entry = {'nombre_archivo': r['nombre_archivo'], 'corporacion': r['corporacion'],
                     'fuente': r['fuente'], 'path_token': _e14_pres_path_token(r['ruta_archivo'])}
            out.setdefault(r['fuente'] or 'transmision', []).append(entry)
        return jsonify({'success': True, 'data': out, 'total': sum(len(v) for v in out.values())})
    except Exception as e:
        logger.exception('[e14-pres/buscar]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/e14-pres/buscar-por-mesa', methods=['GET'])
def e14_pres_buscar_por_mesa():
    """Busca E14 PRES de una mesa específica (resuelve divipol desde idmesa)."""
    err = _require_session()
    if err: return err
    try:
        idmesa = request.args.get('idmesa', type=int)
        if not idmesa:
            return jsonify({'success': False, 'error': 'Falta idmesa'}), 400
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""SELECT d.coddepto, d.codmipio, d.codzona, d.codpuesto, dm.mesa
                               FROM divipolmesa_presidencial_2026 dm
                               JOIN divipol_presidencial_2026 d ON d.iddivipol=dm.iddivipol AND d.clase='P'
                               WHERE dm.idmesa=%s LIMIT 1""", (idmesa,))
                div = cur.fetchone()
                if not div:
                    return jsonify({'success': False, 'error': 'Mesa no encontrada en divipol'})
                cur.execute("""SELECT ruta_archivo, fuente
                               FROM e14_index_presidencial
                               WHERE coddepto=%s AND codmipio=%s AND codzona=%s
                                 AND codpuesto=%s AND mesa=%s
                               ORDER BY CASE fuente WHEN 'claveros' THEN 1 WHEN 'delegados' THEN 2 ELSE 3 END
                               LIMIT 1""",
                            (str(div['coddepto']).zfill(2), str(div['codmipio']).zfill(3),
                             str(div['codzona']).zfill(3), str(div['codpuesto']).zfill(2),
                             str(div['mesa']).zfill(3)))
                e14 = cur.fetchone()
        if not e14:
            return jsonify({'success': False, 'error': 'E14 PRES no encontrado para esta mesa'})
        return jsonify({'success': True,
                        'path_token': _e14_pres_path_token(e14['ruta_archivo']),
                        'fuente': e14['fuente']})
    except Exception as e:
        logger.exception('[e14-pres/buscar-por-mesa]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/e14-pres/ver/<path_token>', methods=['GET'])
def e14_pres_ver(path_token):
    """Sirve un PDF E14 desde el repositorio."""
    err = _require_session()
    if err: return err
    try:
        rel = _e14_pres_decode_token(path_token)
        ruta = os.path.normpath(os.path.join(E14_PRES_BASE_PATH, rel))
        if not ruta.startswith(os.path.normpath(E14_PRES_BASE_PATH)):
            return jsonify({'success': False, 'error': 'Ruta no permitida'}), 403
        if not os.path.isfile(ruta):
            return jsonify({'success': False, 'error': 'Archivo no encontrado'}), 404
        from flask import send_file
        resp = send_file(ruta, mimetype='application/pdf', conditional=True)
        resp.headers['Cache-Control'] = 'public, max-age=86400'
        resp.headers['Accept-Ranges'] = 'bytes'
        return resp
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/e14-pres/pagina-imagen', methods=['GET'])
def e14_pres_pagina_imagen():
    """Extrae una página de un PDF E14 PRES como imagen PNG (base64). Requiere pdftoppm."""
    err = _require_session()
    if err: return err
    try:
        import subprocess, tempfile
        token = request.args.get('path_token', '')
        page = request.args.get('page', type=int)
        if not token or not page:
            return jsonify({'success': False, 'error': 'path_token y page requeridos'}), 400
        rel = _e14_pres_decode_token(token)
        ruta = os.path.normpath(os.path.join(E14_PRES_BASE_PATH, rel))
        if not ruta.startswith(os.path.normpath(E14_PRES_BASE_PATH)):
            return jsonify({'success': False, 'error': 'Ruta no permitida'}), 403
        if not os.path.isfile(ruta):
            return jsonify({'success': False, 'error': 'Archivo no encontrado'}), 404
        with tempfile.TemporaryDirectory() as tmp:
            prefix = os.path.join(tmp, 'page')
            r = subprocess.run(
                ['pdftoppm','-png','-f',str(page),'-l',str(page),'-r','200','-singlefile',ruta,prefix],
                capture_output=True, timeout=30)
            if r.returncode != 0:
                return jsonify({'success': False, 'error': f'pdftoppm: {r.stderr.decode()[:200]}'}), 500
            png = prefix + '.png'
            if not os.path.exists(png):
                return jsonify({'success': False, 'error': 'No se generó imagen'}), 500
            with open(png, 'rb') as f: data = f.read()
        return jsonify({'success': True,
                        'imagen_base64': 'data:image/png;base64,' + _b64.b64encode(data).decode()})
    except Exception as e:
        logger.exception('[e14-pres/pagina-imagen]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/e14-pres/monitor/resumen', methods=['GET'])
def e14_pres_monitor_resumen():
    """Cobertura total nacional de E14 PRES."""
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT
                      COUNT(DISTINCT dm.idmesa) AS total_mesas,
                      COUNT(DISTINCT CASE WHEN e.id IS NOT NULL THEN dm.idmesa END) AS mesas_cubiertas,
                      COUNT(DISTINCT e.id) AS total_e14,
                      COUNT(DISTINCT CASE WHEN e.fuente='claveros'    THEN e.id END) AS claveros,
                      COUNT(DISTINCT CASE WHEN e.fuente='delegados'   THEN e.id END) AS delegados,
                      COUNT(DISTINCT CASE WHEN e.fuente='transmision' THEN e.id END) AS transmision
                    FROM divipol_presidencial_2026 d
                    JOIN divipolmesa_presidencial_2026 dm ON dm.iddivipol = d.iddivipol
                    LEFT JOIN e14_index_presidencial e ON
                        LPAD(d.coddepto::text,2,'0') = e.coddepto AND
                        LPAD(d.codmipio::text,3,'0') = e.codmipio AND
                        LPAD(d.codzona::text,3,'0')  = e.codzona  AND
                        d.codpuesto = e.codpuesto AND
                        LPAD(dm.mesa::text,3,'0') = e.mesa
                    WHERE d.clase='P'
                """)
                row = cur.fetchone()
        tm = row['total_mesas'] or 0
        mc = row['mesas_cubiertas'] or 0
        cobertura = round(mc/tm*100, 2) if tm > 0 else 0
        return jsonify({'success': True, 'data': {
            'total_mesas': tm, 'mesas_cubiertas': mc, 'total_e14': row['total_e14'] or 0,
            'cobertura': cobertura,
            'claveros': row['claveros'] or 0, 'delegados': row['delegados'] or 0,
            'transmision': row['transmision'] or 0,
            'ruta': E14_PRES_BASE_PATH,
            'existe_ruta': os.path.isdir(E14_PRES_BASE_PATH),
        }})
    except Exception as e:
        logger.exception('[e14-pres/monitor/resumen]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/e14-pres/monitor/por-depto', methods=['GET'])
def e14_pres_monitor_depto():
    """Cobertura por departamento."""
    err = _require_session()
    if err: return err
    try:
        fuente = request.args.get('fuente')
        extra, params = '', []
        if fuente:
            extra = ' AND e.fuente = %s'
            params.append(fuente)
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(f"""
                    SELECT
                      LPAD(d.coddepto::text,2,'0') AS coddepto,
                      MAX(d.nomdepto) AS nomdepto,
                      COUNT(DISTINCT dm.idmesa) AS mesas_esperadas,
                      COUNT(DISTINCT CASE WHEN e.id IS NOT NULL THEN dm.idmesa END) AS mesas_cubiertas,
                      COUNT(DISTINCT e.id) AS e14_recibidos,
                      COUNT(DISTINCT CASE WHEN e.fuente='claveros'    THEN e.id END) AS claveros,
                      COUNT(DISTINCT CASE WHEN e.fuente='delegados'   THEN e.id END) AS delegados,
                      COUNT(DISTINCT CASE WHEN e.fuente='transmision' THEN e.id END) AS transmision
                    FROM divipol_presidencial_2026 d
                    JOIN divipolmesa_presidencial_2026 dm ON dm.iddivipol = d.iddivipol
                    LEFT JOIN e14_index_presidencial e ON
                        LPAD(d.coddepto::text,2,'0') = e.coddepto AND
                        LPAD(d.codmipio::text,3,'0') = e.codmipio AND
                        LPAD(d.codzona::text,3,'0')  = e.codzona  AND
                        d.codpuesto = e.codpuesto AND
                        LPAD(dm.mesa::text,3,'0') = e.mesa
                        {extra}
                    WHERE d.clase='P'
                    GROUP BY d.coddepto
                    ORDER BY d.coddepto
                """, params)
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        logger.exception('[e14-pres/monitor/por-depto]')
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== COMISIONES ESCRUTINIO ====================

@app.route('/api/comisiones-pres/cargar-excel', methods=['POST'])
def com_pres_cargar_excel():
    err = _require_session()
    if err: return err
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    if 'archivo' not in request.files:
        return jsonify({'success': False, 'error': 'No se envió archivo'}), 400
    try:
        import openpyxl, io as _io2
        wb = openpyxl.load_workbook(_io2.BytesIO(request.files['archivo'].read()), read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2:
            return jsonify({'success': False, 'error': 'Archivo vacío'}), 400
        data_rows = rows[1:]

        def _si(v):
            if v is None: return None
            try: return int(v)
            except (ValueError, TypeError): return None
        def _ss(v):
            if v is None: return None
            s = str(v).strip()
            return s if s else None

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('TRUNCATE TABLE distribucion_comisiones_presidencial_2026 RESTART IDENTITY')
                sql = '''INSERT INTO distribucion_comisiones_presidencial_2026
                         (coddepto, nomdepto, codmpio, nommpio, zona, codpuesto, nompuesto,
                          codcomuna, nomcomuna, comision_nacional, comision_dptal, comision_municipal,
                          comision_zonal, comision_auxiliar, nombre_comision, mesa_inicial, mesa_final,
                          total_mesas, tipo_comision, lugar_escrutinios, direccion_escrutinios)
                         VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'''
                batch = []
                for r in data_rows:
                    # 21 columnas según el formato
                    if len(r) < 21:
                        r = list(r) + [None] * (21 - len(r))
                    batch.append((
                        _ss(r[0]), _ss(r[1]), _ss(r[2]), _ss(r[3]),
                        _ss(r[4]), _ss(r[5]), _ss(r[6]),
                        _ss(r[7]), _ss(r[8]),
                        _si(r[9]), _si(r[10]), _si(r[11]),
                        _si(r[12]), _si(r[13]), _ss(r[14]),
                        _si(r[15]), _si(r[16]), _si(r[17]),
                        _ss(r[18]), _ss(r[19]), _ss(r[20])
                    ))
                cur.executemany(sql, batch)
                conn.commit()
        return jsonify({'success': True, 'message': f'{len(data_rows):,} registros cargados.'})
    except Exception as e:
        logger.exception('[comisiones-pres/cargar]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/comisiones-pres/resumen', methods=['GET'])
def com_pres_resumen():
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('SELECT COUNT(*) AS total FROM distribucion_comisiones_presidencial_2026')
                total = cur.fetchone()['total']
                cur.execute("""SELECT COUNT(DISTINCT nombre_comision) AS comisiones
                               FROM distribucion_comisiones_presidencial_2026""")
                comisiones = cur.fetchone()['comisiones']
                cur.execute("""SELECT tipo_comision, COUNT(*) AS n, SUM(total_mesas) AS mesas
                               FROM distribucion_comisiones_presidencial_2026
                               WHERE tipo_comision IS NOT NULL
                               GROUP BY tipo_comision ORDER BY n DESC""")
                por_tipo = cur.fetchall()
                cur.execute("""SELECT coddepto, MIN(nomdepto) AS nomdepto,
                                      COUNT(DISTINCT nombre_comision) AS comisiones,
                                      COUNT(*) AS asignaciones,
                                      SUM(total_mesas) AS total_mesas
                               FROM distribucion_comisiones_presidencial_2026
                               WHERE coddepto IS NOT NULL
                               GROUP BY coddepto ORDER BY coddepto""")
                por_depto = cur.fetchall()
        return jsonify({'success': True, 'total': total, 'comisiones': comisiones,
                        'por_tipo': por_tipo, 'por_depto': por_depto})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/comisiones-pres/departamentos', methods=['GET'])
def com_pres_deptos():
    err = _require_session()
    if err: return err
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""SELECT DISTINCT coddepto, UPPER(nomdepto) AS nomdepto
                           FROM distribucion_comisiones_presidencial_2026
                           WHERE coddepto IS NOT NULL ORDER BY coddepto""")
            return jsonify({'success': True, 'data': cur.fetchall()})

@app.route('/api/comisiones-pres/municipios', methods=['GET'])
def com_pres_mpios():
    err = _require_session()
    if err: return err
    cd = request.args.get('coddepto')
    if not cd: return jsonify({'success': True, 'data': []})
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""SELECT DISTINCT codmpio, UPPER(nommpio) AS nommpio
                           FROM distribucion_comisiones_presidencial_2026
                           WHERE coddepto=%s AND codmpio IS NOT NULL ORDER BY nommpio""", (cd,))
            return jsonify({'success': True, 'data': cur.fetchall()})

@app.route('/api/comisiones-pres/zonas', methods=['GET'])
def com_pres_zonas():
    err = _require_session()
    if err: return err
    cd = request.args.get('coddepto'); cm = request.args.get('codmpio')
    where = ['zona IS NOT NULL']; params = []
    if cd: where.append('coddepto=%s'); params.append(cd)
    if cm: where.append('codmpio=%s'); params.append(cm)
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(f"""SELECT DISTINCT zona FROM distribucion_comisiones_presidencial_2026
                            WHERE {' AND '.join(where)} ORDER BY zona""", params)
            return jsonify({'success': True, 'data': cur.fetchall()})

@app.route('/api/comisiones-pres/consultar', methods=['GET'])
def com_pres_consultar():
    err = _require_session()
    if err: return err
    try:
        cd = request.args.get('coddepto', '')
        cm = request.args.get('codmpio', '')
        cz = request.args.get('zona', '')
        tipo = request.args.get('tipo', '')
        nombre = request.args.get('nombre', '')
        pagina = max(1, int(request.args.get('pagina', 1)))
        por_pagina = min(500, max(20, int(request.args.get('por_pagina', 100))))

        where, params = [], []
        if cd:     where.append('coddepto=%s');  params.append(cd)
        if cm:     where.append('codmpio=%s');   params.append(cm)
        if cz:     where.append('zona=%s');      params.append(cz)
        if tipo:   where.append('tipo_comision=%s'); params.append(tipo)
        if nombre: where.append('nombre_comision ILIKE %s'); params.append(f'%{nombre}%')
        where_sql = ' AND '.join(where) if where else '1=1'

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(f"SELECT COUNT(*) AS n FROM distribucion_comisiones_presidencial_2026 WHERE {where_sql}", params)
                total = cur.fetchone()['n']
                offset = (pagina - 1) * por_pagina
                cur.execute(f"""SELECT * FROM distribucion_comisiones_presidencial_2026
                                WHERE {where_sql}
                                ORDER BY coddepto, codmpio, zona, codpuesto, nombre_comision
                                LIMIT %s OFFSET %s""", params + [por_pagina, offset])
                filas = cur.fetchall()
        return jsonify({'success': True, 'data': filas, 'total': total,
                        'pagina': pagina, 'por_pagina': por_pagina,
                        'paginas': (total + por_pagina - 1) // por_pagina})
    except Exception as e:
        logger.exception('[comisiones-pres/consultar]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/comisiones-pres/e24-disponibles', methods=['GET'])
def com_pres_e24_disponibles():
    """Lista códigos de comisión que tienen E24 PRES disponible.
    Formato esperado: E24_PRES_<DEPTO>_<MPIO>_..._<COD_COMISION>.pdf
    """
    err = _require_session()
    if err: return err
    try:
        import re as _re_e
        disp = {}
        if os.path.isdir(E24_PRES_BASE_PATH):
            for f in os.listdir(E24_PRES_BASE_PATH):
                if not (f.lower().endswith('.pdf') and f.startswith('E24_')):
                    continue
                m = _re_e.match(r'E24_(PRES|PRE)_(\d+)_(\d+)_.*_(\d+)\.pdf$', f, _re_e.I)
                if m:
                    cd = m.group(2); cm = m.group(3); ccom = m.group(4)
                    key = f'{cd}_{cm}_{ccom}'
                    disp[key] = f
        return jsonify({'success': True, 'data': disp, 'total': len(disp),
                        'ruta': E24_PRES_BASE_PATH,
                        'existe_ruta': os.path.isdir(E24_PRES_BASE_PATH)})
    except Exception as e:
        logger.exception('[com-pres/e24-disp]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/comisiones-pres/e24/<filename>')
def com_pres_e24_pdf(filename):
    """Sirve un PDF E24 PRES desde el repositorio configurado."""
    err = _require_session()
    if err: return err
    import re as _re_e
    if not _re_e.match(r'^E24_[A-Za-z0-9_]+\.pdf$', filename):
        return jsonify({'success': False, 'error': 'Nombre de archivo inválido'}), 400
    ruta = os.path.normpath(os.path.join(E24_PRES_BASE_PATH, filename))
    if not ruta.startswith(os.path.normpath(E24_PRES_BASE_PATH)):
        return jsonify({'success': False, 'error': 'Ruta inválida'}), 400
    if not os.path.isfile(ruta):
        return jsonify({'success': False, 'error': 'Archivo no encontrado'}), 404
    from flask import send_file
    return send_file(ruta, mimetype='application/pdf')

@app.route('/api/comisiones-pres/vaciar', methods=['POST'])
def com_pres_vaciar():
    err = _require_session()
    if err: return err
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute('TRUNCATE TABLE distribucion_comisiones_presidencial_2026 RESTART IDENTITY')
            conn.commit()
    return jsonify({'success': True, 'message': 'Tabla vaciada'})

# ==================== AGE — Observaciones Acta General Escrutinio ====================
import threading as _threading
import re as _re_age

_age_pres_progreso = {'activo': False, 'total': 0, 'procesados': 0,
                      'registros': 0, 'actual': '', 'errores': [], 'terminado': False}

def _parsear_age_pres(ruta_docx):
    """Parser de .docx AGE — extrae observaciones por mesa."""
    from docx import Document
    doc = Document(ruta_docx)
    obs_list = []
    last = {'coddepto': '', 'nomdepto': '', 'codmpio': '', 'nommpio': ''}
    corp_map = {'01': 'SENADO', '02': 'CAMARA', '03': 'CNS', '06': 'CONSULTAS',
                '001': 'PRESIDENTE', '1': 'SENADO', '2': 'CAMARA'}

    for p in doc.paragraphs:
        txt = p.text.strip()
        if 'DEPARTAMENTO' not in txt:
            continue
        registros = _re_age.split(r'(?=DEPARTAMENTO)', txt)
        for reg in registros:
            reg = reg.strip()
            if not reg or not reg.startswith('DEPARTAMENTO'):
                continue
            depto_m = _re_age.search(r'DEPARTAMENTO\s+(\d+)-([^,]+)', reg)
            if not depto_m:
                m_only = _re_age.search(r'DEPARTAMENTO\s+(\d+)\s+MUNICIPIO', reg)
                if m_only:
                    cd = m_only.group(1).strip()
                    nd = last['nomdepto'] if last['coddepto'].lstrip('0') == cd.lstrip('0') else ''
                    depto_m = type('M', (), {'group': lambda self, n, _cd=cd, _nd=nd: _cd if n==1 else _nd})()
            mpio_m = _re_age.search(r'MUNICIPIO\s+(\d+)-([^,]+)', reg)
            if not mpio_m:
                m_only = _re_age.search(r'MUNICIPIO\s+(\d+)\s+(?:ANTE|ZONA)', reg)
                if m_only:
                    cm = m_only.group(1).strip()
                    nm = last['nommpio'] if last['codmpio'].lstrip('0') == cm.lstrip('0') else ''
                    mpio_m = type('M', (), {'group': lambda self, n, _cm=cm, _nm=nm: _cm if n==1 else _nm})()

            zona_m   = _re_age.search(r'ZONA\s+(\d+)', reg)
            puesto_m = _re_age.search(r'PUESTO\s+(\d+)-([^,]+?)(?=,\s*MESA)', reg)
            mesa_m   = _re_age.search(r'MESA\s+(?:N[°º]\s*)?(\d+)', reg)
            corp_m   = _re_age.search(r'(\d+)-(SENADO|CAMARA|CNS|CONSULTAS|PRESIDENTE)', reg, _re_age.I)
            if corp_m:
                corp_name = corp_m.group(2).upper()
            else:
                cm_code = _re_age.search(r'CORPORACION\s+(\d+)', reg)
                corp_name = corp_map.get(cm_code.group(1), cm_code.group(1)) if cm_code else ''

            if not mesa_m:
                continue

            votos_urna_m = _re_age.search(r'total de votos en la urna[^=]*=\s*(\d+)', reg, _re_age.I)
            votos_inc_m  = _re_age.search(r'total votos incinerados\s*=\s*(\d+)', reg, _re_age.I)
            sufrag_m     = _re_age.search(r'sufragantes E-?11[^=]*=\s*(\d+)', reg, _re_age.I)
            jurados_m    = _re_age.search(r'firmada por\s+(\d+)\s+jurados', reg, _re_age.I)
            fecha_m      = _re_age.search(r'En la fecha\s+([\d-]+\s+[\d:]+\s*[ap]\.\s*m\.)', reg, _re_age.I)

            tiene_tach = 'si tiene tachaduras' in reg.lower() or 'sí tiene tachaduras' in reg.lower()
            tiene_rec  = 'recuento' in reg.lower() and 'no registra' not in reg.lower()

            tipo = 'rutinaria'; obs_texto = ''
            if 'modificaci' in reg.lower():
                tipo = 'modificacion'
            elif tiene_tach:
                spec = _re_age.search(r'borrones o otro,\s*(.+?)(?:,\s*el acta|$)', reg, _re_age.I)
                if spec and len(spec.group(1).strip()) > 5:
                    tipo = 'tachadura_con_detalle'; obs_texto = spec.group(1).strip()
                else:
                    tipo = 'tachadura'
            obs_mesa = _re_age.search(r'observaci[oó]n de la mesa\s+(.+?)(?:\.\s*Esta informacion|$)', reg, _re_age.I)
            if obs_mesa:
                obs_texto = obs_mesa.group(1).strip()
                if tipo == 'rutinaria':
                    tipo = 'observacion_mesa'
            if tipo == 'modificacion':
                mod = _re_age.search(r'modificaci[oó]n de la mesa con la siguiente informaci[oó]n,?\s*(.+)', reg, _re_age.I)
                if mod:
                    obs_texto = mod.group(1).strip()

            # Partido
            nompartido = ''
            for pat in [
                r'(?:EN\s+EL\s+)?PARTIDO\s+(?:DE\s+LA\s+)?(CONSERVADOR\s+COLOMBIANO|LIBERAL\s+COLOMBIANO|CAMBIO\s+RADICAL|CENTRO\s+DEMOCR[AÁ]TICO|VERDE|ALIANZA\s+VERDE|POLO\s+DEMOCR[AÁ]TICO|COLOMBIA\s+HUMANA|PACTO\s+HIST[OÓ]RICO|SALVACI[OÓ]N\s+NACIONAL|MIRA|COMUNES|ASI|MAIS|U\b)',
                r'EN\s+EL\s+PARTIDO\s+([A-ZÁÉÍÓÚÑ\s]+?)(?:\s+EL|\s+SE|\s+TOTAL|\s+CAT)',
            ]:
                pm = _re_age.search(pat, obs_texto or reg, _re_age.I)
                if pm: nompartido = pm.group(1).strip().rstrip('.'); break

            # Memorizar para herencia
            if depto_m:
                last['coddepto'] = depto_m.group(1).strip()
                last['nomdepto'] = depto_m.group(2).strip()
            if mpio_m:
                last['codmpio'] = mpio_m.group(1).strip()
                last['nommpio'] = mpio_m.group(2).strip()

            if tipo == 'rutinaria':
                continue

            obs_list.append({
                'coddepto': depto_m.group(1).strip() if depto_m else '',
                'nomdepto': depto_m.group(2).strip() if depto_m else '',
                'codmpio':  mpio_m.group(1).strip() if mpio_m else '',
                'nommpio':  mpio_m.group(2).strip() if mpio_m else '',
                'zona':      zona_m.group(1).strip() if zona_m else '',
                'codpuesto': puesto_m.group(1).strip() if puesto_m else '',
                'nompuesto': puesto_m.group(2).strip() if puesto_m else '',
                'mesa':      mesa_m.group(1).strip(),
                'corporacion': corp_name,
                'codpartido': '', 'nompartido': nompartido,
                'codcandidato': '', 'nomcandidato': '',
                'tipo_observacion': tipo,
                'observacion': obs_texto[:2000] if obs_texto else '',
                'votos_urna':        int(votos_urna_m.group(1)) if votos_urna_m else None,
                'votos_incinerados': int(votos_inc_m.group(1)) if votos_inc_m else None,
                'sufragantes_e11':   int(sufrag_m.group(1))    if sufrag_m else None,
                'jurados_firma':     int(jurados_m.group(1))   if jurados_m else None,
                'tiene_tachaduras': tiene_tach,
                'tiene_recuento':   tiene_rec,
                'fecha_registro':   fecha_m.group(1) if fecha_m else '',
            })
    return obs_list

def _is_admin():
    return session.get('id_perfil') == 1

@app.route('/api/age-pres/escanear', methods=['POST'])
def age_pres_escanear():
    err = _require_session()
    if err: return err
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Acceso denegado'}), 403
    try:
        d = request.get_json() or {}
        carpeta = (d.get('carpeta') or '').strip()
        if not carpeta or not os.path.isdir(carpeta):
            return jsonify({'success': False, 'error': f'Carpeta no encontrada: {carpeta}'})
        archivos = [f for f in os.listdir(carpeta) if f.lower().endswith('.docx') and not f.startswith('~$')]
        return jsonify({'success': True, 'archivos': len(archivos), 'carpeta': carpeta})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/age-pres/cargar-carpeta', methods=['POST'])
def age_pres_cargar_carpeta():
    global _age_pres_progreso
    err = _require_session()
    if err: return err
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Acceso denegado'}), 403
    if _age_pres_progreso['activo']:
        return jsonify({'success': False, 'error': 'Ya hay una carga en curso'})
    try:
        d = request.get_json() or {}
        carpeta = (d.get('carpeta') or '').strip()
        procesados_dir = (d.get('procesados') or '').strip()
        if not carpeta or not os.path.isdir(carpeta):
            return jsonify({'success': False, 'error': f'Carpeta no encontrada: {carpeta}'})
        if procesados_dir:
            os.makedirs(procesados_dir, exist_ok=True)
        archivos = sorted([f for f in os.listdir(carpeta)
                           if f.lower().endswith('.docx') and not f.startswith('~$')])
        if not archivos:
            return jsonify({'success': False, 'error': 'No hay archivos .docx en la carpeta'})

        _age_pres_progreso = {'activo': True, 'total': len(archivos), 'procesados': 0,
                              'registros': 0, 'actual': '', 'errores': [], 'terminado': False}

        def procesar():
            global _age_pres_progreso
            import shutil as _sh
            for nombre in archivos:
                _age_pres_progreso['actual'] = nombre
                ruta = os.path.join(carpeta, nombre)
                try:
                    obs = _parsear_age_pres(ruta)
                    if obs:
                        with get_db_connection() as conn:
                            with conn.cursor() as cur:
                                cur.execute('DELETE FROM age_presidencial_2026 WHERE archivo = %s', (nombre,))
                                cur.executemany("""
                                    INSERT INTO age_presidencial_2026
                                    (coddepto, nomdepto, codmpio, nommpio, zona, codpuesto, nompuesto, mesa,
                                     corporacion, codpartido, nompartido, codcandidato, nomcandidato,
                                     tipo_observacion, observacion, votos_urna, votos_incinerados,
                                     sufragantes_e11, jurados_firma, tiene_tachaduras, tiene_recuento,
                                     fecha_registro, archivo)
                                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                                """, [(o['coddepto'], o['nomdepto'], o['codmpio'], o['nommpio'],
                                       o['zona'], o['codpuesto'], o['nompuesto'], o['mesa'],
                                       o['corporacion'], o['codpartido'], o['nompartido'],
                                       o['codcandidato'], o['nomcandidato'], o['tipo_observacion'],
                                       o['observacion'], o['votos_urna'], o['votos_incinerados'],
                                       o['sufragantes_e11'], o['jurados_firma'], o['tiene_tachaduras'],
                                       o['tiene_recuento'], o['fecha_registro'], nombre) for o in obs])
                                conn.commit()
                        _age_pres_progreso['registros'] += len(obs)
                    if procesados_dir:
                        dst = os.path.join(procesados_dir, nombre)
                        if os.path.exists(dst): os.remove(dst)
                        _sh.move(ruta, dst)
                except Exception as ex:
                    _age_pres_progreso['errores'].append(f'{nombre}: {ex}')
                _age_pres_progreso['procesados'] += 1
            _age_pres_progreso['activo'] = False
            _age_pres_progreso['terminado'] = True
            _age_pres_progreso['actual'] = ''

        _threading.Thread(target=procesar, daemon=True).start()
        return jsonify({'success': True, 'total': len(archivos), 'msg': 'Carga iniciada'})
    except Exception as e:
        _age_pres_progreso['activo'] = False
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/age-pres/progreso', methods=['GET'])
def age_pres_progreso_g():
    err = _require_session()
    if err: return err
    return jsonify({'success': True, **_age_pres_progreso})

@app.route('/api/age-pres/resumen', methods=['GET'])
def age_pres_resumen():
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('SELECT COUNT(*) AS total FROM age_presidencial_2026')
                total = cur.fetchone()['total']
                cur.execute('SELECT COUNT(DISTINCT archivo) AS archivos FROM age_presidencial_2026')
                archivos = cur.fetchone()['archivos']
                cur.execute("""SELECT tipo_observacion, COUNT(*) AS cantidad
                               FROM age_presidencial_2026
                               GROUP BY tipo_observacion ORDER BY cantidad DESC""")
                por_tipo = cur.fetchall()
                cur.execute("""SELECT archivo, COUNT(*) AS observaciones,
                                      COUNT(DISTINCT coddepto || codmpio) AS municipios,
                                      MIN(fecha_carga) AS fecha
                               FROM age_presidencial_2026
                               GROUP BY archivo ORDER BY fecha DESC""")
                por_arch = cur.fetchall()
        return jsonify({'success': True, 'total': total, 'archivos': archivos,
                        'por_tipo': por_tipo, 'por_archivo': por_arch})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/age-pres/filtros', methods=['GET'])
def age_pres_filtros():
    err = _require_session()
    if err: return err
    try:
        depto = request.args.get('depto', '')
        mpio  = request.args.get('mpio', '')
        zona  = request.args.get('zona', '')
        depto_i = int(depto) if depto else None
        mpio_i  = int(mpio)  if mpio  else None
        zona_i  = int(zona)  if zona  else None
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""SELECT LPAD(coddepto::text,2,'0') AS coddepto, nomdepto
                               FROM divipol_presidencial_2026
                               WHERE clase='D' AND coddepto > 0 ORDER BY coddepto""")
                deptos = cur.fetchall()
                mpios = []
                if depto_i is not None:
                    cur.execute("""SELECT LPAD(codmipio::text,3,'0') AS codmpio, nommipio AS nommpio
                                   FROM divipol_presidencial_2026
                                   WHERE clase='M' AND coddepto=%s ORDER BY codmipio""", (depto_i,))
                    mpios = cur.fetchall()
                zonas = []
                if depto_i is not None and mpio_i is not None:
                    cur.execute("""SELECT DISTINCT LPAD(codzona::text,2,'0') AS zona
                                   FROM divipol_presidencial_2026
                                   WHERE clase='Z' AND coddepto=%s AND codmipio=%s ORDER BY zona""",
                                (depto_i, mpio_i))
                    zonas = [r['zona'] for r in cur.fetchall()]
                puestos = []
                if depto_i is not None and mpio_i is not None and zona_i is not None:
                    cur.execute("""SELECT codpuesto, nompuesto
                                   FROM divipol_presidencial_2026
                                   WHERE clase='P' AND coddepto=%s AND codmipio=%s AND codzona=%s
                                   ORDER BY codpuesto""", (depto_i, mpio_i, zona_i))
                    puestos = cur.fetchall()
                cur.execute("SELECT codpartido, nompartido FROM partidos_presidencial_2026 ORDER BY nompartido")
                partidos = cur.fetchall()
                cur.execute("""SELECT DISTINCT tipo_observacion AS tipo FROM age_presidencial_2026
                               WHERE tipo_observacion IS NOT NULL ORDER BY tipo""")
                tipos = [r['tipo'] for r in cur.fetchall()]
        return jsonify({'success': True, 'deptos': deptos, 'mpios': mpios,
                        'zonas': zonas, 'puestos': puestos, 'partidos': partidos,
                        'tipos': tipos})
    except Exception as e:
        logger.exception('[age-pres/filtros]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/age-pres/consultar', methods=['GET'])
def age_pres_consultar():
    err = _require_session()
    if err: return err
    try:
        depto = request.args.get('depto', '')
        mpio  = request.args.get('mpio', '')
        zona  = request.args.get('zona', '')
        puesto= request.args.get('puesto', '')
        mesa  = request.args.get('mesa', '')
        corp  = request.args.get('corporacion', '')
        tipo  = request.args.get('tipo', '')
        partido = request.args.get('partido', '')
        pagina = max(1, int(request.args.get('pagina', 1)))
        por_pagina = min(500, max(20, int(request.args.get('por_pagina', 100))))

        where, params = [], []
        if depto:   where.append("LPAD(coddepto,2,'0')=%s"); params.append(depto.zfill(2))
        if mpio:    where.append("LPAD(codmpio,3,'0')=%s"); params.append(mpio.zfill(3))
        if zona:    where.append("LPAD(zona,2,'0')=%s");    params.append(zona.zfill(2))
        if puesto:  where.append("codpuesto=%s");           params.append(puesto)
        if mesa:    where.append("mesa=%s");                params.append(mesa)
        if corp:    where.append("corporacion=%s");         params.append(corp)
        if tipo:    where.append("tipo_observacion=%s");    params.append(tipo)
        if partido: where.append("nompartido ILIKE %s");    params.append(f'%{partido}%')
        where_sql = ' AND '.join(where) if where else '1=1'

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(f"SELECT COUNT(*) AS n FROM age_presidencial_2026 WHERE {where_sql}", params)
                total = cur.fetchone()['n']
                offset = (pagina - 1) * por_pagina
                cur.execute(f"""SELECT id, coddepto, nomdepto, codmpio, nommpio, zona,
                                       codpuesto, nompuesto, mesa, corporacion, nompartido,
                                       tipo_observacion, observacion, votos_urna, votos_incinerados,
                                       sufragantes_e11, jurados_firma, tiene_tachaduras, tiene_recuento,
                                       fecha_registro, archivo, fecha_carga
                                FROM age_presidencial_2026
                                WHERE {where_sql}
                                ORDER BY coddepto, codmpio, zona, codpuesto, mesa, id
                                LIMIT %s OFFSET %s""", params + [por_pagina, offset])
                filas = cur.fetchall()
        return jsonify({'success': True, 'data': filas, 'total': total,
                        'pagina': pagina, 'por_pagina': por_pagina,
                        'paginas': (total + por_pagina - 1) // por_pagina})
    except Exception as e:
        logger.exception('[age-pres/consultar]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/age-pres/vaciar', methods=['POST'])
def age_pres_vaciar():
    err = _require_session()
    if err: return err
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Acceso denegado'}), 403
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('TRUNCATE TABLE age_presidencial_2026 RESTART IDENTITY')
                conn.commit()
        return jsonify({'success': True, 'message': 'Tabla AGE vaciada'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== EVIDENCIAS (vista de abogados — read-only) ====================
import zipfile, io as _io

@app.route('/api/evidencias-pres/exportar-excel-listado', methods=['GET'])
def evpres_excel_listado():
    """Excel del listado de investigaciones agrupadas."""
    err = _require_session()
    if err: return err
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT i.coddepto, MIN(i.nomdepto) AS nomdepto,
                           i.codcandidato1, MIN(i.nom_candidato1) AS nom_candidato1, MIN(i.nom_partido1) AS nom_partido1,
                           i.codcandidato2, MIN(i.nom_candidato2) AS nom_candidato2, MIN(i.nom_partido2) AS nom_partido2,
                           COUNT(*) AS num_mesas,
                           COUNT(DISTINCT i.codmipio) AS num_mpios,
                           SUM(CASE WHEN EXISTS (
                               SELECT 1 FROM evidencias_presidencial_2026 e
                               WHERE e.idmesa=i.idmesa
                                 AND e.codcandidato IN (i.codcandidato1, i.codcandidato2)
                                 AND COALESCE(e.tipo_formulario,'') NOT IN ('NO_E14','SIN_EVIDENCIA')
                           ) THEN 1 ELSE 0 END) AS mesas_con_evidencia
                    FROM investigaciones_presidencial_2026 i
                    GROUP BY i.coddepto, i.codcandidato1, i.codcandidato2
                    ORDER BY i.coddepto, num_mesas DESC
                """)
                rows = cur.fetchall()

        wb = Workbook(); ws = wb.active; ws.title = 'Investigaciones'
        hf = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        hfill = PatternFill('solid', fgColor='1E3A8A')
        bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        ws.merge_cells('A1:J1')
        ws['A1'] = 'Investigaciones Presidencial 2026 — Listado'
        ws['A1'].font = Font(name='Arial', bold=True, size=13, color='1E3A8A')
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = ['#', 'Cód Depto', 'Departamento',
                   'Cand A', 'Candidato A', 'Partido A',
                   'Cand B', 'Candidato B', 'Mesas', 'Con Evidencia']
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal='center'); c.border = bdr
        for idx, r in enumerate(rows, 4):
            vals = [idx-3, r['coddepto'], r['nomdepto'] or '',
                    r['codcandidato1'], r['nom_candidato1'] or '', r['nom_partido1'] or '',
                    r['codcandidato2'], r['nom_candidato2'] or '',
                    r['num_mesas'] or 0, r['mesas_con_evidencia'] or 0]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=idx, column=ci, value=v); cell.border = bdr
                if ci in (9, 10): cell.number_format = '#,##0'
        widths = [5, 9, 22, 7, 28, 28, 7, 28, 10, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A4'

        out = _io.BytesIO(); wb.save(out); out.seek(0)
        from flask import send_file
        return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'Investigaciones_Presidencial_2026_{date.today().isoformat()}.xlsx')
    except Exception as e:
        logger.exception('[evpres/excel-listado]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/evidencias-pres/exportar-excel-grupo', methods=['GET'])
def evpres_excel_grupo():
    """Excel detallado de un grupo (par candidatos × depto) con todas sus mesas + evidencias."""
    err = _require_session()
    if err: return err
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        coddepto = request.args.get('coddepto', type=int)
        ccd1 = request.args.get('codcandidato1', type=int)
        ccd2 = request.args.get('codcandidato2', type=int)
        if not ccd1 or not ccd2:
            return jsonify({'success': False, 'error': 'codcandidato1 y codcandidato2 requeridos'}), 400

        where = ['i.codcandidato1=%s', 'i.codcandidato2=%s']
        params = [ccd1, ccd2]
        if coddepto is not None:
            where.append('i.coddepto=%s'); params.append(coddepto)

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(f"""
                    SELECT i.idmesa, i.coddepto, i.nomdepto, i.codmipio, i.nommipio,
                           i.codzona, i.codpuesto, i.nompuesto, i.mesa,
                           i.nom_candidato1, i.preconteo1, i.dia_valor1, i.diferencia1,
                           i.nom_candidato2, i.preconteo2, i.dia_valor2, i.diferencia2,
                           i.estado_reclamacion, i.usuario_asignado,
                           (SELECT COUNT(*) FROM evidencias_presidencial_2026 e
                             WHERE e.idmesa=i.idmesa
                               AND e.codcandidato IN (i.codcandidato1, i.codcandidato2)
                               AND COALESCE(e.tipo_formulario,'') NOT IN ('NO_E14','SIN_EVIDENCIA')) AS num_evidencias
                    FROM investigaciones_presidencial_2026 i
                    WHERE {' AND '.join(where)}
                    ORDER BY i.nommipio, i.codzona, i.codpuesto, i.mesa
                """, params)
                rows = cur.fetchall()

        wb = Workbook(); ws = wb.active; ws.title = 'Detalle'
        hf = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        hfill = PatternFill('solid', fgColor='1E3A8A')
        bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        ws.merge_cells('A1:O1')
        nom1 = rows[0]['nom_candidato1'] if rows else f'Cand {ccd1}'
        nom2 = rows[0]['nom_candidato2'] if rows else f'Cand {ccd2}'
        ws['A1'] = f'Investigación Presidencial — {nom1} vs {nom2}'
        ws['A1'].font = Font(name='Arial', bold=True, size=12, color='1E3A8A')
        ws['A1'].alignment = Alignment(horizontal='center')

        headers = ['#','Depto','Municipio','Zona','Puesto','Mesa',
                   'Prec A','Esc A','Dif A','Prec B','Esc B','Dif B',
                   'Estado','Asignado','Evidencias']
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal='center'); c.border = bdr
        for idx, r in enumerate(rows, 4):
            puesto = f"{r['codpuesto'] or ''} - {r['nompuesto'] or ''}"
            vals = [idx-3, r['nomdepto'] or '', r['nommipio'] or '', r['codzona'], puesto, r['mesa'],
                    r['preconteo1'] or 0, r['dia_valor1'] or 0, r['diferencia1'] or 0,
                    r['preconteo2'] or 0, r['dia_valor2'] or 0, r['diferencia2'] or 0,
                    r['estado_reclamacion'] or '', r['usuario_asignado'] or '',
                    r['num_evidencias'] or 0]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=idx, column=ci, value=v); cell.border = bdr
                if ci in (7,8,9,10,11,12,15): cell.number_format = '#,##0'
        widths = [5,16,18,6,30,7,8,8,8,8,8,8,12,14,11]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A4'

        out = _io.BytesIO(); wb.save(out); out.seek(0)
        from flask import send_file
        nombre = f"Detalle_{nom1[:20]}_vs_{nom2[:20]}.xlsx".replace(' ', '_').replace('/', '-')
        return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=nombre)
    except Exception as e:
        logger.exception('[evpres/excel-grupo]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/evidencias-pres/exportar-zip-grupo', methods=['GET'])
def evpres_zip_grupo():
    """ZIP con todas las evidencias físicas (archivos) de un grupo."""
    err = _require_session()
    if err: return err
    try:
        coddepto = request.args.get('coddepto', type=int)
        ccd1 = request.args.get('codcandidato1', type=int)
        ccd2 = request.args.get('codcandidato2', type=int)
        if not ccd1 or not ccd2:
            return jsonify({'success': False, 'error': 'codcandidato1 y codcandidato2 requeridos'}), 400

        where = ['i.codcandidato1=%s', 'i.codcandidato2=%s']
        params = [ccd1, ccd2]
        if coddepto is not None:
            where.append('i.coddepto=%s'); params.append(coddepto)

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(f"""
                    SELECT i.idmesa, i.nomdepto, i.nommipio, i.mesa,
                           e.id AS evid_id, e.nombre_archivo, e.ruta_archivo, e.tipo_formulario,
                           e.observacion, e.codcandidato, e.usuario, e.fecha
                    FROM investigaciones_presidencial_2026 i
                    JOIN evidencias_presidencial_2026 e
                      ON e.idmesa = i.idmesa
                     AND e.codcandidato IN (i.codcandidato1, i.codcandidato2)
                     AND COALESCE(e.tipo_formulario,'') NOT IN ('NO_E14','SIN_EVIDENCIA')
                    WHERE {' AND '.join(where)}
                    ORDER BY i.nomdepto, i.nommipio, i.mesa, e.id
                """, params)
                evs = cur.fetchall()

        if not evs:
            return jsonify({'success': False, 'error': 'No hay evidencias físicas en este grupo'}), 404

        buf = _io.BytesIO()
        archivos_included = 0
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Index Excel mínimo
            idx_lines = ['Mesa\tCandidato\tTipo\tArchivo\tUsuario\tFecha\tObservacion']
            for e in evs:
                depto = (e['nomdepto'] or 'SIN_DEPTO').replace('/', '-').strip()
                mpio = (e['nommipio'] or 'SIN_MPIO').replace('/', '-').strip()
                mesa = str(e['mesa'] or e['idmesa']).zfill(4)
                tipo = e['tipo_formulario'] or 'OTRO'
                nom_arch = e['nombre_archivo'] or f'evidencia_{e["evid_id"]}'
                zip_path = f"{depto}/{mpio}/Mesa_{mesa}/{tipo}_{nom_arch}"
                if e['ruta_archivo'] and os.path.exists(e['ruta_archivo']):
                    zf.write(e['ruta_archivo'], zip_path)
                    archivos_included += 1
                idx_lines.append(
                    f"{mesa}\t{e['codcandidato']}\t{tipo}\t{nom_arch}\t"
                    f"{e['usuario'] or ''}\t{e['fecha'] or ''}\t{(e['observacion'] or '').replace(chr(9),' ')[:200]}"
                )
            zf.writestr('_indice.tsv', '\n'.join(idx_lines))

        if archivos_included == 0:
            return jsonify({'success': False,
                            'error': 'Ningún archivo físico disponible en el servidor (solo metadatos)'}), 404

        buf.seek(0)
        from flask import send_file
        return send_file(buf, mimetype='application/zip', as_attachment=True,
                         download_name=f'evidencias_pres_{ccd1}_vs_{ccd2}.zip')
    except Exception as e:
        logger.exception('[evpres/zip-grupo]')
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== CONSULTA VOTACIÓN ESCRUTINIO ====================

@app.route('/api/escrutinio/dias-procesados', methods=['GET'])
def escrutinio_dias_procesados():
    """Lista los días con seguimiento poblado (numdia + facceso)."""
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT numdia, facceso, procesado, fecha
                    FROM dias_escrutinio_presidencial
                    ORDER BY numdia
                """)
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/escrutinio/repoblar-seguimiento/<facceso>', methods=['POST'])
def escrutinio_repoblar_seguimiento(facceso):
    """Vuelve a invocar fn_poblar_escrutinio_presidencial para una facceso."""
    err = _require_session()
    if err: return err
    msg = _poblar_seguimiento_dia(facceso)
    return jsonify({'success': True, 'message': msg})

def _ultimo_dia_seguimiento():
    """Devuelve el último numdia con datos en seguimiento (o 0)."""
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COALESCE(MAX(numdia),0) AS m FROM dias_escrutinio_presidencial WHERE procesado=TRUE")
            return cur.fetchone()['m']

@app.route('/api/consulta-escrutinio/filtros/departamentos', methods=['GET'])
def consulta_esc_deptos():
    err = _require_session()
    if err: return err
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT coddepto, nomdepto FROM divipol_presidencial_2026 WHERE clase='D' ORDER BY coddepto")
            return jsonify({'success': True, 'data': cur.fetchall()})

@app.route('/api/consulta-escrutinio/filtros/municipios', methods=['GET'])
def consulta_esc_mpios():
    err = _require_session()
    if err: return err
    coddepto = request.args.get('coddepto', type=int)
    if coddepto is None: return jsonify({'success': True, 'data': []})
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT codmipio, nommipio FROM divipol_presidencial_2026 WHERE clase='M' AND coddepto=%s ORDER BY codmipio", (coddepto,))
            return jsonify({'success': True, 'data': cur.fetchall()})

@app.route('/api/consulta-escrutinio/filtros/zonas', methods=['GET'])
def consulta_esc_zonas():
    err = _require_session()
    if err: return err
    cd = request.args.get('coddepto', type=int)
    cm = request.args.get('codmipio', type=int)
    if cd is None or cm is None: return jsonify({'success': True, 'data': []})
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT codzona FROM divipol_presidencial_2026 WHERE clase='Z' AND coddepto=%s AND codmipio=%s ORDER BY codzona", (cd, cm))
            return jsonify({'success': True, 'data': cur.fetchall()})

@app.route('/api/consulta-escrutinio/filtros/puestos', methods=['GET'])
def consulta_esc_puestos():
    err = _require_session()
    if err: return err
    cd = request.args.get('coddepto', type=int)
    cm = request.args.get('codmipio', type=int)
    cz = request.args.get('codzona', type=int)
    if cd is None or cm is None: return jsonify({'success': True, 'data': []})
    sql = "SELECT codpuesto, nompuesto FROM divipol_presidencial_2026 WHERE clase='P' AND coddepto=%s AND codmipio=%s"
    params = [cd, cm]
    if cz is not None:
        sql += ' AND codzona=%s'; params.append(cz)
    sql += ' ORDER BY codpuesto'
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            return jsonify({'success': True, 'data': cur.fetchall()})

@app.route('/api/consulta-escrutinio/filtros/candidatos', methods=['GET'])
def consulta_esc_candidatos():
    err = _require_session()
    if err: return err
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT codcandidato, nomcandidato, codpartido,
                       (SELECT nompartido FROM partidos_presidencial_2026 WHERE codpartido=c.codpartido LIMIT 1) AS nompartido
                FROM candidatos_presidencial_2026 c
                ORDER BY num_tarjeton, codcandidato
            """)
            return jsonify({'success': True, 'data': cur.fetchall()})

def _build_ultimo_valor_expr(numdia_max):
    """Construye COALESCE(diaN, diaN-1, ..., dia1) para tomar el último día válido."""
    if numdia_max <= 0:
        return '0'
    cols = [f'dia{i}' for i in range(numdia_max, 0, -1)]
    return f"COALESCE({','.join(cols)}, 0)"

@app.route('/api/consulta-escrutinio/consultar', methods=['POST'])
def consulta_esc_consultar():
    """Consulta votación por escrutinio: agrega según filtros. Lee de seguimiento_escrutinio_presidencial_2026."""
    err = _require_session()
    if err: return err
    try:
        d = request.get_json() or {}
        cd  = d.get('coddepto')
        cm  = d.get('codmipio')
        cz  = d.get('codzona')
        cp  = d.get('codpuesto')
        ccd = d.get('codcandidato')
        numdia = d.get('numdia')   # día específico, opcional. Si None → último

        ultimo = _ultimo_dia_seguimiento()
        if ultimo == 0:
            return jsonify({'success': True, 'data': [], 'tipo_resultado': 'sin_datos',
                            'message': 'No hay días de escrutinio procesados.'})
        if numdia is None or numdia == '':
            numdia_use = ultimo
        else:
            numdia_use = int(numdia)

        ultimo_expr = _build_ultimo_valor_expr(numdia_use)
        nomcand_sql = _nomcandidato_sql('s')

        # Join: seguimiento → divipol/mesa para geo, candidatos para nombres
        base_from = f"""seguimiento_escrutinio_presidencial_2026 s
            JOIN divipolmesa_presidencial_2026 dm ON dm.idmesa = s.idmesa
            JOIN divipol_presidencial_2026 dv ON dv.iddivipol = dm.iddivipol AND dv.clase = 'P'
            LEFT JOIN candidatos_presidencial_2026 c
                   ON c.codcandidato = s.codcandidato AND c.codpartido = s.codpartido
            LEFT JOIN partidos_presidencial_2026 pt
                   ON pt.codpartido = s.codpartido"""

        where, params = [], []
        if cd  is not None: where.append('dv.coddepto = %s');     params.append(int(cd))
        if cm  is not None: where.append('dv.codmipio = %s');     params.append(int(cm))
        if cz  is not None: where.append('dv.codzona = %s');      params.append(int(cz))
        if cp  is not None: where.append('dv.codpuesto = %s');    params.append(str(cp).zfill(2))
        if ccd is not None: where.append('s.codcandidato = %s');  params.append(int(ccd))
        where.append(f'({ultimo_expr}) > 0')
        where_sql = ' AND '.join(where)

        sin_filtros = not any([cd, cm, cz, cp, ccd])
        solo_cand   = ccd is not None and not any([cd, cm, cz, cp])
        depto_only  = cd is not None and not any([cm, cz, cp])
        mpio_only   = cd is not None and cm is not None and not any([cz, cp])
        zona_only   = cd is not None and cm is not None and cz is not None and cp is None
        puesto_only = cd is not None and cm is not None and cz is not None and cp is not None

        if sin_filtros or solo_cand:
            sql = f"""SELECT s.codcandidato,
                            {nomcand_sql} AS nomcandidato,
                            COALESCE(pt.nompartido, '—') AS nompartido,
                            SUM({ultimo_expr}) AS total_votos
                     FROM {base_from}
                     WHERE {where_sql}
                     GROUP BY s.codcandidato, c.nomcandidato, pt.nompartido
                     ORDER BY total_votos DESC LIMIT 500"""
            tipo = 'resumen_nacional' if sin_filtros else 'candidato_nacional'
        elif depto_only or mpio_only:
            sql = f"""SELECT s.codcandidato,
                            {nomcand_sql} AS nomcandidato,
                            COALESCE(pt.nompartido, '—') AS nompartido,
                            SUM({ultimo_expr}) AS total_votos
                     FROM {base_from}
                     WHERE {where_sql}
                     GROUP BY s.codcandidato, c.nomcandidato, pt.nompartido
                     ORDER BY total_votos DESC LIMIT 500"""
            tipo = 'por_departamento' if depto_only else 'por_municipio'
        elif zona_only:
            sql = f"""SELECT dv.codpuesto, dv.nompuesto, s.codcandidato,
                            {nomcand_sql} AS nomcandidato,
                            SUM({ultimo_expr}) AS total_votos
                     FROM {base_from}
                     WHERE {where_sql}
                     GROUP BY dv.codpuesto, dv.nompuesto, s.codcandidato, c.nomcandidato
                     ORDER BY dv.codpuesto, total_votos DESC LIMIT 1000"""
            tipo = 'por_zona'
        elif puesto_only:
            sql = f"""SELECT dm.mesa, s.codcandidato,
                            {nomcand_sql} AS nomcandidato,
                            COALESCE(pt.nompartido, '—') AS nompartido,
                            SUM({ultimo_expr}) AS total_votos
                     FROM {base_from}
                     WHERE {where_sql}
                     GROUP BY dm.mesa, s.codcandidato, c.nomcandidato, pt.nompartido
                     ORDER BY dm.mesa, total_votos DESC LIMIT 5000"""
            tipo = 'por_puesto_mesas'
        else:
            sql = f"""SELECT dv.coddepto, dv.codmipio, dv.codzona, dv.codpuesto, dm.mesa,
                            s.codcandidato,
                            {nomcand_sql} AS nomcandidato,
                            SUM({ultimo_expr}) AS total_votos
                     FROM {base_from}
                     WHERE {where_sql}
                     GROUP BY dv.coddepto, dv.codmipio, dv.codzona, dv.codpuesto, dm.mesa,
                              s.codcandidato, c.nomcandidato
                     ORDER BY total_votos DESC LIMIT 500"""
            tipo = 'detalle'

        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(sql, params)
                rows = cur.fetchall()

        return jsonify({'success': True, 'tipo_resultado': tipo, 'data': rows,
                        'total': len(rows), 'numdia': numdia_use, 'ultimo_dia': ultimo})
    except Exception as e:
        logger.exception('[consulta-escrutinio/consultar]')
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== HEALTHCHECK ====================
@app.route('/api/health')
def health():
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute('SELECT 1 AS ok')
                cur.fetchone()
        return jsonify({'success': True, 'db': 'ok', 'app': 'auditor-presidencial-2026'})
    except Exception as e:
        return jsonify({'success': False, 'db': 'error', 'error': str(e)}), 500

# ==================== GENERADOR INCREMENTAL E14 ====================
import hashlib, threading as _threading, re as _re_gen
from werkzeug.utils import secure_filename as _secure

GEN_DIR_PLANTILLAS = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'plantillas')
GEN_DIR_CORTES     = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'generated')
os.makedirs(GEN_DIR_PLANTILLAS, exist_ok=True)
os.makedirs(GEN_DIR_CORTES, exist_ok=True)

_gen_stop = _threading.Event()
_gen_thread = None
_gen_thread_lock = _threading.Lock()

def _gen_norm(s):
    if s is None: return ''
    s = str(s).strip().upper()
    for a, b in (('Á','A'),('É','E'),('Í','I'),('Ó','O'),('Ú','U'),('Ü','U'),('Ñ','N')):
        s = s.replace(a, b)
    s = _re_gen.sub(r'[^A-Z0-9 ]', ' ', s)
    s = _re_gen.sub(r'\s+', ' ', s).strip()
    return s

def _gen_extraer_alias_candidato(header):
    """De '#1 CEPEDA CASTRO' devuelve ('CEPEDA CASTRO', 1) — alias y posible cod_tarjeton."""
    if not header: return None, None
    m = _re_gen.match(r'\s*#?\s*(\d+)\s+(.*)$', str(header).strip())
    if m:
        return _gen_norm(m.group(2)), int(m.group(1))
    return _gen_norm(header), None

def _gen_parsear_plantilla(ruta):
    """Lee el XLSX y devuelve (mesas_raw, candidatos_raw). NO toca BD."""
    import openpyxl
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 6:
        raise ValueError('Plantilla sin filas suficientes')
    # Buscar fila de headers: primera fila cuyo col 0 sea exactamente "MESA ID"
    fila_hdr = None
    for i, r in enumerate(rows[:15]):
        if r and r[0] and _gen_norm(r[0]) == 'MESA ID':
            fila_hdr = i; break
    if fila_hdr is None:
        raise ValueError('No encontré la fila de encabezados (Mesa ID)')
    hdr = rows[fila_hdr]
    # Localizar columnas
    cols = {'mesa_id':0, 'depto':None, 'mpio':None, 'puesto':None, 'mesa':None,
            'blanco':None, 'nulo':None, 'nomarc':None, 'total':None}
    candidatos = []  # [(col_idx, header_text, alias, cod_tarjeton)]
    for j, h in enumerate(hdr):
        if h is None: continue
        hs = _gen_norm(h)
        if hs == 'MESA ID' or 'MESA ID' in hs: cols['mesa_id'] = j
        elif hs in ('DEPARTAMENTO','DEPTO'): cols['depto'] = j
        elif hs in ('MUNICIPIO','MPIO','MUNICIPIO MESA'): cols['mpio'] = j
        elif hs in ('PUESTO',): cols['puesto'] = j
        elif hs == 'MESA': cols['mesa'] = j
        elif 'BLANCO' in hs: cols['blanco'] = j
        elif 'NULO' in hs: cols['nulo'] = j
        elif 'NO MARCADO' in hs or 'NOMARCADO' in hs: cols['nomarc'] = j
        elif 'TOTAL' in hs: cols['total'] = j
        elif _re_gen.match(r'\s*#?\s*\d+\s+', str(h)):
            alias, cod = _gen_extraer_alias_candidato(h)
            candidatos.append((j, str(h).strip(), alias, cod))
    if not all([cols['depto'] is not None, cols['mpio'] is not None,
                cols['puesto'] is not None, cols['mesa'] is not None]):
        raise ValueError(f'Faltan columnas requeridas: cols={cols}')
    if not candidatos:
        raise ValueError('No detecté columnas de candidatos (formato esperado: "#N APELLIDO")')

    mesas_raw = []
    for i in range(fila_hdr + 1, len(rows)):
        r = rows[i]
        if not r or not r[cols['mesa_id']]: continue
        mid = str(r[cols['mesa_id']]).strip()
        if mid.upper() == 'TOTALES': break
        mesas_raw.append({
            'fila': i + 1,  # 1-indexed para openpyxl
            'mesa_id': mid,
            'depto': r[cols['depto']],
            'mpio': r[cols['mpio']],
            'puesto': r[cols['puesto']],
            'mesa': r[cols['mesa']],
        })
    return {
        'mesas': mesas_raw,
        'candidatos': candidatos,
        'cols': cols,
    }

def _gen_resolver_mesas(cur, mesas_raw):
    """Cruza nombres con divipol_presidencial_2026 para obtener (idmesa, coddepto, ...)."""
    cur.execute("""
        SELECT dm.idmesa, d.coddepto, d.codmipio, d.codzona, d.codpuesto,
               UPPER(d.nomdepto) AS nd, UPPER(d.nommipio) AS nm, UPPER(d.nompuesto) AS np,
               LPAD(dm.mesa::text,3,'0') AS mesa_str,
               dm.mesa AS mesa_num
        FROM divipol_presidencial_2026 d
        JOIN divipolmesa_presidencial_2026 dm ON dm.iddivipol = d.iddivipol
        WHERE d.clase='P'
    """)
    mapa = {}
    for row in cur.fetchall():
        k = (_gen_norm(row['nd']), _gen_norm(row['nm']), _gen_norm(row['np']), row['mesa_str'])
        mapa[k] = (row['idmesa'], row['coddepto'], row['codmipio'],
                   row['codzona'], row['codpuesto'], row['mesa_num'])
    resueltas = []
    for m in mesas_raw:
        mesa_str = str(m['mesa'] or '').strip().zfill(3)
        k = (_gen_norm(m['depto']), _gen_norm(m['mpio']), _gen_norm(m['puesto']), mesa_str)
        hit = mapa.get(k)
        out = dict(m)
        out['mesa_norm'] = mesa_str
        if hit:
            out['idmesa'], out['coddepto'], out['codmipio'], out['codzona'], out['codpuesto'], out['mesa_int'] = hit
        else:
            out['idmesa'] = None
            out['coddepto'] = out['codmipio'] = out['codzona'] = None
            out['codpuesto'] = None; out['mesa_int'] = None
        resueltas.append(out)
    return resueltas

def _gen_resolver_candidatos(cur, candidatos_raw):
    """Mapea cada columna de candidato al codcandidato real por apellido(s)."""
    cur.execute("SELECT codcandidato, nomcandidato, num_tarjeton FROM candidatos_presidencial_2026")
    cands = [(c['codcandidato'], _gen_norm(c['nomcandidato']), c['num_tarjeton']) for c in cur.fetchall()]
    out = []
    for col_idx, header, alias, cod_tarj in candidatos_raw:
        match = None
        # 1) por num_tarjeton si está
        if cod_tarj is not None:
            for cc, _, nt in cands:
                if nt is not None and int(nt) == cod_tarj:
                    match = cc; break
        # 2) por todas las palabras del alias dentro de nomcandidato
        if not match and alias:
            tokens = [t for t in alias.split() if len(t) >= 3]
            for cc, nom, _ in cands:
                if tokens and all(t in nom for t in tokens):
                    match = cc; break
        # 3) por la primera palabra del alias
        if not match and alias:
            first = alias.split()[0] if alias.split() else ''
            if len(first) >= 4:
                for cc, nom, _ in cands:
                    if first in nom:
                        match = cc; break
        out.append({'columna_xlsx': col_idx, 'header_text': header, 'alias': alias, 'codcandidato': match})
    return out

def _gen_estado_get(cur):
    cur.execute("SELECT * FROM gen_estado WHERE id=1")
    return cur.fetchone()

def _gen_get_plantilla_activa(cur):
    cur.execute("""SELECT * FROM gen_plantilla_e14
                   WHERE activa=TRUE
                   ORDER BY fecha_carga DESC LIMIT 1""")
    return cur.fetchone()

@app.route('/api/generador/plantilla', methods=['POST'])
def gen_subir_plantilla():
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    if 'archivo' not in request.files:
        return jsonify({'success': False, 'error': 'Falta archivo'}), 400
    f = request.files['archivo']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'error': 'Solo se acepta .xlsx'}), 400
    safe = _secure(f.filename)
    ts = _time_mod.strftime('%Y%m%d_%H%M%S', _time_mod.localtime())
    nombre_archivo = f'{ts}_{safe}'
    ruta = os.path.join(GEN_DIR_PLANTILLAS, nombre_archivo)
    f.save(ruta)
    try:
        parsed = _gen_parsear_plantilla(ruta)
    except Exception as e:
        os.unlink(ruta)
        return jsonify({'success': False, 'error': f'Plantilla inválida: {e}'}), 400
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE gen_plantilla_e14 SET activa=FALSE WHERE activa=TRUE")
                cur.execute("""INSERT INTO gen_plantilla_e14
                               (nombre, ruta_servidor, mesas_total, candidatos_total, usuario)
                               VALUES (%s,%s,%s,%s,%s) RETURNING id""",
                            (f.filename, ruta, len(parsed['mesas']),
                             len(parsed['candidatos']), session.get('cedula') or ''))
                pid = cur.fetchone()['id']
                mesas_res = _gen_resolver_mesas(cur, parsed['mesas'])
                cur.executemany("""INSERT INTO gen_plantilla_mesas
                                   (plantilla_id, fila_xlsx, mesa_id_str, nomdepto, nommipio,
                                    nompuesto, mesa_num, coddepto, codmipio, codzona, codpuesto, mesa, idmesa)
                                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                                [(pid, m['fila'], m['mesa_id'], m['depto'], m['mpio'],
                                  m['puesto'], m['mesa_norm'], m['coddepto'], m['codmipio'],
                                  m['codzona'], m['codpuesto'], m['mesa_int'], m['idmesa'])
                                 for m in mesas_res])
                cands_res = _gen_resolver_candidatos(cur, parsed['candidatos'])
                cur.executemany("""INSERT INTO gen_plantilla_candidatos
                                   (plantilla_id, columna_xlsx, header_text, alias, codcandidato)
                                   VALUES (%s,%s,%s,%s,%s)""",
                                [(pid, c['columna_xlsx'], c['header_text'],
                                  c['alias'], c['codcandidato']) for c in cands_res])
                cur.execute("UPDATE gen_estado SET plantilla_id=%s WHERE id=1", (pid,))
                conn.commit()
                mesas_ok = sum(1 for m in mesas_res if m['idmesa'])
                cands_ok = sum(1 for c in cands_res if c['codcandidato'])
        return jsonify({'success': True, 'data': {
            'plantilla_id': pid, 'mesas_total': len(mesas_res), 'mesas_resueltas': mesas_ok,
            'candidatos_total': len(cands_res), 'candidatos_resueltos': cands_ok
        }})
    except Exception as e:
        logger.exception('[gen/plantilla]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/generador/plantilla/preview', methods=['GET'])
def gen_plantilla_preview():
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                p = _gen_get_plantilla_activa(cur)
                if not p:
                    return jsonify({'success': True, 'data': None})
                cur.execute("""SELECT fila_xlsx, mesa_id_str, nomdepto, nommipio, nompuesto,
                                       mesa_num, idmesa
                               FROM gen_plantilla_mesas WHERE plantilla_id=%s
                               ORDER BY fila_xlsx""", (p['id'],))
                mesas = cur.fetchall()
                cur.execute("""SELECT columna_xlsx, header_text, alias, codcandidato,
                                      (SELECT nomcandidato FROM candidatos_presidencial_2026
                                       WHERE codcandidato=g.codcandidato) AS nomcandidato
                               FROM gen_plantilla_candidatos g
                               WHERE plantilla_id=%s ORDER BY columna_xlsx""", (p['id'],))
                cands = cur.fetchall()
        return jsonify({'success': True, 'data': {
            'plantilla': dict(p),
            'mesas': mesas,
            'candidatos': cands,
            'mesas_resueltas': sum(1 for m in mesas if m['idmesa']),
            'candidatos_resueltos': sum(1 for c in cands if c['codcandidato']),
        }})
    except Exception as e:
        logger.exception('[gen/plantilla/preview]')
        return jsonify({'success': False, 'error': str(e)}), 500

def _gen_snapshot_y_hash(cur, plantilla_id):
    """Trae los votos actuales del preconteo para las mesas de la plantilla. Devuelve (datos, hash, total_votos, mesas_reportadas)."""
    cur.execute("""
        SELECT pm.fila_xlsx, p.codcandidato, SUM(p.votos)::BIGINT AS votos
        FROM gen_plantilla_mesas pm
        JOIN preconteo_presidencial_2026 p
          ON p.coddepto = pm.coddepto AND p.codmipio = pm.codmipio
         AND p.codzona  = pm.codzona  AND p.codpuesto = pm.codpuesto
         AND p.mesa     = pm.mesa
        WHERE pm.plantilla_id=%s AND pm.idmesa IS NOT NULL
        GROUP BY pm.fila_xlsx, p.codcandidato
    """, (plantilla_id,))
    datos = {}  # fila_xlsx -> {codcand: votos}
    total = 0
    for r in cur.fetchall():
        datos.setdefault(r['fila_xlsx'], {})[r['codcandidato']] = int(r['votos'] or 0)
        total += int(r['votos'] or 0)
    # hash determinista
    seria = ''
    for fila in sorted(datos):
        for cc in sorted(datos[fila]):
            seria += f'{fila}|{cc}|{datos[fila][cc]};'
    h = hashlib.md5(seria.encode()).hexdigest() if seria else 'EMPTY'
    return datos, h, total, len(datos)

def _gen_construir_excel(plantilla_ruta, datos, mapeo_cols_cand, fila_a_geo,
                          plantilla_nombre, num_corte, mesas_rep, total_votos, ts_str):
    """Carga plantilla y rellena las filas que tienen datos."""
    import openpyxl
    wb = openpyxl.load_workbook(plantilla_ruta)
    ws = wb[wb.sheetnames[0]]
    # Buscar columnas especiales por header
    cols_blanco = cols_nulo = cols_nomarc = cols_total = None
    fila_hdr = None
    for r in range(1, min(ws.max_row, 15) + 1):
        v = ws.cell(row=r, column=1).value
        if v and _gen_norm(v) == 'MESA ID':
            fila_hdr = r; break
    if fila_hdr:
        for c in range(1, ws.max_column + 1):
            h = _gen_norm(ws.cell(row=fila_hdr, column=c).value)
            if 'BLANCO' in h: cols_blanco = c
            elif 'NULO' in h: cols_nulo = c
            elif 'NO MARCADO' in h or 'NOMARCADO' in h: cols_nomarc = c
            elif h == 'TOTAL' or 'TOTAL' in h: cols_total = c
    # Llenar filas con datos
    for fila, votos_cand in datos.items():
        suma = 0
        for col_xlsx, codcand in mapeo_cols_cand.items():
            if codcand and codcand in votos_cand:
                v = votos_cand[codcand]
                ws.cell(row=fila, column=col_xlsx + 1).value = v  # col_xlsx 0-indexed
                suma += v
        if cols_blanco and 996 in votos_cand:
            ws.cell(row=fila, column=cols_blanco).value = votos_cand[996]; suma += votos_cand[996]
        if cols_nulo and 997 in votos_cand:
            ws.cell(row=fila, column=cols_nulo).value = votos_cand[997]; suma += votos_cand[997]
        if cols_nomarc and 998 in votos_cand:
            ws.cell(row=fila, column=cols_nomarc).value = votos_cand[998]; suma += votos_cand[998]
        if cols_total:
            ws.cell(row=fila, column=cols_total).value = suma
    # Anotar el subtítulo (fila 2 si parece subtítulo)
    try:
        cell_a2 = ws.cell(row=2, column=1)
        cell_a2.value = f'Corte #{num_corte} | {ts_str} | {mesas_rep} mesas reportadas | {total_votos:,} votos'
    except Exception:
        pass
    base = os.path.splitext(plantilla_nombre)[0]
    nombre_out = f'{base}_corte_{num_corte:04d}_{ts_str.replace(":","").replace(" ","_").replace("-","")}.xlsx'
    ruta_out = os.path.join(GEN_DIR_CORTES, nombre_out)
    wb.save(ruta_out)
    return nombre_out, ruta_out

def _gen_ejecutar_corte():
    """Ejecuta un corte. Devuelve dict con resultado."""
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            est = _gen_estado_get(cur)
            if not est:
                return {'ok': False, 'error': 'Sin estado'}
            p = _gen_get_plantilla_activa(cur)
            if not p:
                return {'ok': False, 'error': 'Sin plantilla activa'}
            datos, h, total_v, mesas_rep = _gen_snapshot_y_hash(cur, p['id'])
            num_corte = (est['ultimo_corte_num'] or 0) + 1
            ts_local = _time_mod.localtime()
            ts_str = _time_mod.strftime('%Y-%m-%d %H:%M:%S', ts_local)
            ts_fname = _time_mod.strftime('%Y%m%d_%H%M%S', ts_local)
            # ¿Sin cambios?
            if est['ultimo_hash'] and est['ultimo_hash'] == h:
                cur.execute("""INSERT INTO gen_cortes
                               (plantilla_id, num_corte, tipo, mesas_reportadas, total_votos, hash_snapshot)
                               VALUES (%s,%s,'sin_cambios',%s,%s,%s)""",
                            (p['id'], num_corte, mesas_rep, total_v, h))
                cur.execute("""UPDATE gen_estado
                               SET ultimo_corte_num=%s, ultimo_corte_at=NOW(),
                                   skipped_consecutivos=skipped_consecutivos+1
                               WHERE id=1""", (num_corte,))
                conn.commit()
                return {'ok': True, 'tipo': 'sin_cambios', 'num_corte': num_corte,
                        'mesas_reportadas': mesas_rep, 'total_votos': total_v}
            # Generar archivo
            cur.execute("""SELECT columna_xlsx, codcandidato FROM gen_plantilla_candidatos
                           WHERE plantilla_id=%s""", (p['id'],))
            mapeo_cols = {r['columna_xlsx']: r['codcandidato'] for r in cur.fetchall()}
            cur.execute("""SELECT fila_xlsx, coddepto, codmipio, codzona, codpuesto, mesa
                           FROM gen_plantilla_mesas WHERE plantilla_id=%s""", (p['id'],))
            geo = {r['fila_xlsx']: r for r in cur.fetchall()}
            nombre, ruta = _gen_construir_excel(
                p['ruta_servidor'], datos, mapeo_cols, geo,
                p['nombre'], num_corte, mesas_rep, total_v, ts_str)
            cur.execute("""INSERT INTO gen_cortes
                           (plantilla_id, num_corte, tipo, archivo, ruta,
                            mesas_reportadas, total_votos, hash_snapshot)
                           VALUES (%s,%s,'generado',%s,%s,%s,%s,%s)""",
                        (p['id'], num_corte, nombre, ruta, mesas_rep, total_v, h))
            cur.execute("""UPDATE gen_estado
                           SET ultimo_corte_num=%s, ultimo_corte_at=NOW(),
                               ultimo_hash=%s, skipped_consecutivos=0
                           WHERE id=1""", (num_corte, h))
            conn.commit()
            return {'ok': True, 'tipo': 'generado', 'num_corte': num_corte, 'archivo': nombre,
                    'mesas_reportadas': mesas_rep, 'total_votos': total_v}

def _gen_loop():
    """Hilo background: ejecuta cortes en bucle."""
    logger.info('[gen] hilo iniciado')
    while not _gen_stop.is_set():
        try:
            res = _gen_ejecutar_corte()
            logger.info(f'[gen] corte: {res}')
        except Exception as e:
            logger.exception('[gen] error en corte')
        # leer intervalo dinámicamente
        try:
            with get_db_connection() as conn:
                with conn.cursor() as cur:
                    est = _gen_estado_get(cur)
                    if not est or not est['activo']:
                        logger.info('[gen] estado=inactivo, saliendo del loop')
                        break
                    intervalo = max(1, int(est['intervalo_min'] or 5))
        except Exception:
            intervalo = 5
        _gen_stop.wait(timeout=intervalo * 60)
    logger.info('[gen] hilo terminado')

@app.route('/api/generador/iniciar', methods=['POST'])
def gen_iniciar():
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    d = request.get_json(silent=True) or {}
    intervalo = int(d.get('intervalo_min') or 5)
    intervalo = max(1, min(60, intervalo))
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                p = _gen_get_plantilla_activa(cur)
                if not p:
                    return jsonify({'success': False, 'error': 'Cargue una plantilla primero'}), 400
                cur.execute("""UPDATE gen_estado
                               SET activo=TRUE, intervalo_min=%s, inicio_at=NOW(),
                                   plantilla_id=%s
                               WHERE id=1""", (intervalo, p['id']))
                conn.commit()
        with _gen_thread_lock:
            global _gen_thread
            if _gen_thread and _gen_thread.is_alive():
                _gen_stop.set()
                _gen_thread.join(timeout=2)
            _gen_stop.clear()
            _gen_thread = _threading.Thread(target=_gen_loop, daemon=True)
            _gen_thread.start()
        return jsonify({'success': True, 'data': {'intervalo_min': intervalo}})
    except Exception as e:
        logger.exception('[gen/iniciar]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/generador/detener', methods=['POST'])
def gen_detener():
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE gen_estado SET activo=FALSE WHERE id=1")
                conn.commit()
        _gen_stop.set()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/generador/estado', methods=['GET'])
def gen_estado():
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                est = _gen_estado_get(cur)
                p = _gen_get_plantilla_activa(cur)
                if p:
                    cur.execute("""SELECT COUNT(*) AS total,
                                          SUM(CASE WHEN idmesa IS NOT NULL THEN 1 ELSE 0 END) AS ok
                                   FROM gen_plantilla_mesas WHERE plantilla_id=%s""", (p['id'],))
                    mres = cur.fetchone()
                    cur.execute("""SELECT COUNT(*) AS total,
                                          SUM(CASE WHEN codcandidato IS NOT NULL THEN 1 ELSE 0 END) AS ok
                                   FROM gen_plantilla_candidatos WHERE plantilla_id=%s""", (p['id'],))
                    cres = cur.fetchone()
                cur.execute("""SELECT COUNT(*) AS gen, COUNT(*) FILTER (WHERE tipo='generado') AS reales,
                                      COUNT(*) FILTER (WHERE tipo='sin_cambios') AS saltados
                               FROM gen_cortes""")
                tot = cur.fetchone()
        thread_alive = bool(_gen_thread and _gen_thread.is_alive())
        return jsonify({'success': True, 'data': {
            'estado': dict(est) if est else None,
            'thread_alive': thread_alive,
            'plantilla': dict(p) if p else None,
            'mesas_resueltas': (mres['ok'] if p else 0), 'mesas_total': (mres['total'] if p else 0),
            'candidatos_resueltos': (cres['ok'] if p else 0), 'candidatos_total': (cres['total'] if p else 0),
            'cortes_total': tot['gen'] or 0,
            'cortes_reales': tot['reales'] or 0,
            'cortes_saltados': tot['saltados'] or 0,
        }})
    except Exception as e:
        logger.exception('[gen/estado]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/generador/cortes', methods=['GET'])
def gen_cortes_lista():
    err = _require_session()
    if err: return err
    try:
        limit = int(request.args.get('limit', 100))
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""SELECT id, num_corte, tipo, archivo, mesas_reportadas,
                                       total_votos, fecha
                               FROM gen_cortes ORDER BY num_corte DESC LIMIT %s""", (limit,))
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/generador/cortes/<int:cid>/descargar', methods=['GET'])
def gen_corte_descargar(cid):
    err = _require_session()
    if err: return err
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT ruta, archivo, tipo FROM gen_cortes WHERE id=%s", (cid,))
                row = cur.fetchone()
        if not row or row['tipo'] != 'generado':
            return jsonify({'success': False, 'error': 'Corte sin archivo'}), 404
        if not os.path.isfile(row['ruta']):
            return jsonify({'success': False, 'error': 'Archivo no encontrado en servidor'}), 404
        from flask import send_file
        return send_file(row['ruta'], as_attachment=True, download_name=row['archivo'])
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/generador/ejecutar-ahora', methods=['POST'])
def gen_ejecutar_manual():
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    try:
        res = _gen_ejecutar_corte()
        return jsonify({'success': True, 'data': res})
    except Exception as e:
        logger.exception('[gen/ejecutar-ahora]')
        return jsonify({'success': False, 'error': str(e)}), 500

def _gen_arrancar_si_activo():
    """Llamado en boot: si gen_estado.activo=TRUE, levantar el hilo."""
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                est = _gen_estado_get(cur)
                if est and est['activo']:
                    global _gen_thread
                    _gen_stop.clear()
                    _gen_thread = _threading.Thread(target=_gen_loop, daemon=True)
                    _gen_thread.start()
                    logger.info('[gen] reanudado al arranque')
    except Exception:
        logger.exception('[gen] no pude arrancar')

_gen_arrancar_si_activo()

# ==================== ADMINISTRACIÓN: USUARIOS Y PERFILES ====================

@app.route('/api/admin/perfiles', methods=['GET'])
def admin_perfiles_listar():
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""SELECT p.id, p.nombre,
                                      (SELECT COUNT(*) FROM usuarios u WHERE u.id_perfil=p.id) AS usuarios
                               FROM perfiles p ORDER BY p.id""")
                return jsonify({'success': True, 'data': cur.fetchall()})
    except Exception as e:
        logger.exception('[admin/perfiles]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/perfiles', methods=['POST'])
def admin_perfil_crear():
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    d = request.get_json(silent=True) or {}
    nombre = (d.get('nombre') or '').strip()
    if not nombre:
        return jsonify({'success': False, 'error': 'Nombre requerido'}), 400
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("INSERT INTO perfiles(nombre) VALUES (%s) RETURNING id", (nombre,))
                pid = cur.fetchone()['id']
                conn.commit()
        return jsonify({'success': True, 'data': {'id': pid, 'nombre': nombre}})
    except Exception as e:
        if 'duplicate' in str(e).lower() or 'unique' in str(e).lower():
            return jsonify({'success': False, 'error': 'Ya existe un perfil con ese nombre'}), 400
        logger.exception('[admin/perfil/crear]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/perfiles/<int:pid>', methods=['PUT'])
def admin_perfil_editar(pid):
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    d = request.get_json(silent=True) or {}
    nombre = (d.get('nombre') or '').strip()
    if not nombre:
        return jsonify({'success': False, 'error': 'Nombre requerido'}), 400
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("UPDATE perfiles SET nombre=%s WHERE id=%s", (nombre, pid))
                if cur.rowcount == 0:
                    return jsonify({'success': False, 'error': 'Perfil no encontrado'}), 404
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        if 'duplicate' in str(e).lower():
            return jsonify({'success': False, 'error': 'Ya existe un perfil con ese nombre'}), 400
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/perfiles/<int:pid>', methods=['DELETE'])
def admin_perfil_eliminar(pid):
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) AS n FROM usuarios WHERE id_perfil=%s", (pid,))
                n = cur.fetchone()['n']
                if n > 0:
                    return jsonify({'success': False, 'error': f'No se puede eliminar: {n} usuario(s) asignado(s)'}), 400
                cur.execute("DELETE FROM perfiles WHERE id=%s", (pid,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/usuarios', methods=['GET'])
def admin_usuarios_listar():
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    try:
        q = (request.args.get('q') or '').strip()
        id_perfil = request.args.get('id_perfil', type=int)
        page = max(1, request.args.get('page', type=int) or 1)
        per_page = min(200, request.args.get('per_page', type=int) or 50)
        where = []
        params = []
        if q:
            where.append("(u.cedula ILIKE %s OR u.nombres ILIKE %s OR u.apellidos ILIKE %s OR u.correo ILIKE %s)")
            patt = f'%{q}%'
            params.extend([patt, patt, patt, patt])
        if id_perfil:
            where.append("u.id_perfil = %s")
            params.append(id_perfil)
        wsql = ('WHERE ' + ' AND '.join(where)) if where else ''
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(f"SELECT COUNT(*) AS n FROM usuarios u {wsql}", params)
                total = cur.fetchone()['n']
                cur.execute(f"""
                    SELECT u.id, u.cedula, u.nombres, u.apellidos, u.correo,
                           u.id_perfil, p.nombre AS perfil, u.creado
                    FROM usuarios u
                    LEFT JOIN perfiles p ON p.id = u.id_perfil
                    {wsql}
                    ORDER BY u.id DESC
                    LIMIT %s OFFSET %s
                """, params + [per_page, (page - 1) * per_page])
                rows = cur.fetchall()
        return jsonify({'success': True, 'data': rows,
                        'total': total, 'page': page, 'per_page': per_page})
    except Exception as e:
        logger.exception('[admin/usuarios]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/usuarios', methods=['POST'])
def admin_usuario_crear():
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    d = request.get_json(silent=True) or {}
    cedula = (d.get('cedula') or '').strip()
    contrasena = d.get('contrasena') or ''
    if not cedula or not contrasena:
        return jsonify({'success': False, 'error': 'Cédula y contraseña requeridas'}), 400
    if len(contrasena) < 4:
        return jsonify({'success': False, 'error': 'Contraseña mínimo 4 caracteres'}), 400
    id_perfil = d.get('id_perfil')
    if not id_perfil:
        return jsonify({'success': False, 'error': 'Perfil requerido'}), 400
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""INSERT INTO usuarios(cedula, contrasena, nombres, apellidos, correo, id_perfil)
                               VALUES (%s,%s,%s,%s,%s,%s) RETURNING id""",
                            (cedula, hash_password(contrasena),
                             (d.get('nombres') or '').strip(),
                             (d.get('apellidos') or '').strip(),
                             (d.get('correo') or '').strip() or None,
                             int(id_perfil)))
                uid = cur.fetchone()['id']
                conn.commit()
        return jsonify({'success': True, 'data': {'id': uid}})
    except Exception as e:
        if 'duplicate' in str(e).lower() or 'unique' in str(e).lower():
            return jsonify({'success': False, 'error': 'Ya existe un usuario con esa cédula'}), 400
        logger.exception('[admin/usuario/crear]')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/usuarios/<int:uid>', methods=['PUT'])
def admin_usuario_editar(uid):
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    d = request.get_json(silent=True) or {}
    campos, params = [], []
    for k in ('cedula', 'nombres', 'apellidos', 'correo'):
        if k in d:
            v = (d.get(k) or '').strip()
            campos.append(f'{k} = %s'); params.append(v or None)
    if 'id_perfil' in d and d['id_perfil']:
        # No permitir bajar de admin al propio usuario
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id_perfil FROM usuarios WHERE id=%s", (uid,))
                row = cur.fetchone()
                if row and row['id_perfil'] == 1 and int(d['id_perfil']) != 1 and uid == session.get('user_id'):
                    return jsonify({'success': False, 'error': 'No puedes cambiar tu propio rol de admin'}), 400
        campos.append('id_perfil = %s'); params.append(int(d['id_perfil']))
    if not campos:
        return jsonify({'success': False, 'error': 'Nada que actualizar'}), 400
    params.append(uid)
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(f"UPDATE usuarios SET {', '.join(campos)} WHERE id=%s", params)
                if cur.rowcount == 0:
                    return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        if 'duplicate' in str(e).lower():
            return jsonify({'success': False, 'error': 'Cédula duplicada'}), 400
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/usuarios/<int:uid>/password', methods=['POST'])
def admin_usuario_reset_pwd(uid):
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    d = request.get_json(silent=True) or {}
    pwd = d.get('contrasena') or ''
    if len(pwd) < 4:
        return jsonify({'success': False, 'error': 'Contraseña mínimo 4 caracteres'}), 400
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("""UPDATE usuarios SET contrasena=%s, session_token=NULL
                               WHERE id=%s""", (hash_password(pwd), uid))
                if cur.rowcount == 0:
                    return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/usuarios/<int:uid>', methods=['DELETE'])
def admin_usuario_eliminar(uid):
    if not _is_admin():
        return jsonify({'success': False, 'error': 'Solo admin'}), 403
    if uid == session.get('user_id'):
        return jsonify({'success': False, 'error': 'No puedes eliminar tu propio usuario'}), 400
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id_perfil FROM usuarios WHERE id=%s", (uid,))
                row = cur.fetchone()
                if not row:
                    return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404
                if row['id_perfil'] == 1:
                    cur.execute("SELECT COUNT(*) AS n FROM usuarios WHERE id_perfil=1")
                    if cur.fetchone()['n'] <= 1:
                        return jsonify({'success': False, 'error': 'No puedes eliminar al último administrador'}), 400
                cur.execute("DELETE FROM usuarios WHERE id=%s", (uid,))
                conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== DASHBOARD MÉTRICAS ====================
@app.route('/api/dashboard/metricas', methods=['GET'])
def dashboard_metricas():
    """Métricas globales para el dashboard: mesas, cobertura, evidencias, AGE, comisiones, top deptos y candidatos."""
    err = _require_session()
    if err: return err
    try:
        ultimo = _ultimo_dia_seguimiento()
        expr = _build_ultimo_valor_expr(ultimo) if ultimo > 0 else '0'
        out = {}
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                # ---- Mesas universo ----
                cur.execute("SELECT COUNT(*) AS n FROM divipolmesa_presidencial_2026")
                out['total_mesas'] = cur.fetchone()['n'] or 0

                # ---- Preconteo: mesas distintas con votos ----
                cur.execute("""SELECT COUNT(*) AS n FROM (
                                 SELECT DISTINCT coddepto, codmipio, codzona, codpuesto, mesa
                                 FROM preconteo_presidencial_2026 WHERE votos > 0
                               ) t""")
                out['mesas_preconteo'] = cur.fetchone()['n'] or 0

                # ---- Escrutinio: mesas con al menos un día reportado ----
                if ultimo > 0:
                    cur.execute(f"""SELECT COUNT(DISTINCT idmesa) AS n
                                    FROM seguimiento_escrutinio_presidencial_2026
                                    WHERE {expr} > 0""")
                    out['mesas_escrutinio'] = cur.fetchone()['n'] or 0
                else:
                    out['mesas_escrutinio'] = 0
                out['ultimo_dia'] = ultimo

                # ---- Evidencias ----
                cur.execute("""SELECT
                                 COUNT(*) AS total,
                                 COUNT(DISTINCT idmesa) AS mesas,
                                 COUNT(*) FILTER (WHERE tipo_formulario='E14') AS e14,
                                 COUNT(*) FILTER (WHERE tipo_formulario='E24') AS e24,
                                 COUNT(*) FILTER (WHERE tipo_formulario='NO_E14') AS no_e14,
                                 COUNT(*) FILTER (WHERE tipo_formulario='SIN_EVIDENCIA') AS sin_ev,
                                 COUNT(*) FILTER (WHERE tipo_formulario NOT IN ('E14','E24','NO_E14','SIN_EVIDENCIA')
                                                  OR tipo_formulario IS NULL) AS otros
                               FROM evidencias_presidencial_2026""")
                r = cur.fetchone()
                out['evidencias'] = {
                    'total': r['total'] or 0, 'mesas': r['mesas'] or 0,
                    'e14': r['e14'] or 0, 'e24': r['e24'] or 0,
                    'no_e14': r['no_e14'] or 0, 'sin_ev': r['sin_ev'] or 0, 'otros': r['otros'] or 0
                }

                # ---- Investigaciones ----
                cur.execute("""SELECT
                                 COUNT(*) AS total,
                                 COUNT(*) FILTER (WHERE estado_reclamacion='pendiente') AS pendientes,
                                 COUNT(*) FILTER (WHERE estado_reclamacion='en_proceso') AS en_proceso,
                                 COUNT(*) FILTER (WHERE estado_reclamacion='resuelta') AS resueltas,
                                 COUNT(DISTINCT idmesa) AS mesas
                               FROM investigaciones_presidencial_2026""")
                r = cur.fetchone()
                out['investigaciones'] = {
                    'total': r['total'] or 0, 'pendientes': r['pendientes'] or 0,
                    'en_proceso': r['en_proceso'] or 0, 'resueltas': r['resueltas'] or 0,
                    'mesas': r['mesas'] or 0
                }

                # ---- AGE ----
                cur.execute("""SELECT
                                 COUNT(*) AS total,
                                 COUNT(DISTINCT (coddepto, codmpio, zona, codpuesto, mesa)) AS mesas,
                                 COUNT(*) FILTER (WHERE tiene_tachaduras = TRUE) AS tachaduras,
                                 COUNT(*) FILTER (WHERE tiene_recuento = TRUE) AS recuento
                               FROM age_presidencial_2026""")
                r = cur.fetchone()
                out['age'] = {
                    'total': r['total'] or 0, 'mesas': r['mesas'] or 0,
                    'tachaduras': r['tachaduras'] or 0, 'recuento': r['recuento'] or 0
                }

                # ---- Comisiones escrutadoras ----
                cur.execute("""SELECT
                                 COUNT(DISTINCT nombre_comision) AS total,
                                 COUNT(DISTINCT tipo_comision) AS tipos
                               FROM distribucion_comisiones_presidencial_2026""")
                r = cur.fetchone()
                out['comisiones'] = {'total': r['total'] or 0, 'tipos': r['tipos'] or 0}
                cur.execute("""SELECT tipo_comision, COUNT(DISTINCT nombre_comision) AS c
                               FROM distribucion_comisiones_presidencial_2026
                               WHERE tipo_comision IS NOT NULL
                               GROUP BY tipo_comision ORDER BY c DESC""")
                out['comisiones_tipos'] = cur.fetchall()

                # ---- E14 indexados ----
                cur.execute("""SELECT
                                 COUNT(*) AS total,
                                 COUNT(*) FILTER (WHERE fuente='claveros') AS claveros,
                                 COUNT(*) FILTER (WHERE fuente='delegados') AS delegados,
                                 COUNT(*) FILTER (WHERE fuente='transmision') AS transmision,
                                 COUNT(DISTINCT (coddepto, codmipio, codzona, codpuesto, mesa)) AS mesas
                               FROM e14_index_presidencial""")
                r = cur.fetchone()
                out['e14'] = {
                    'total': r['total'] or 0, 'mesas': r['mesas'] or 0,
                    'claveros': r['claveros'] or 0, 'delegados': r['delegados'] or 0,
                    'transmision': r['transmision'] or 0
                }

                # ---- Catálogo ----
                cur.execute("SELECT COUNT(*) AS n FROM candidatos_presidencial_2026")
                out['candidatos'] = cur.fetchone()['n'] or 0
                cur.execute("SELECT COUNT(*) AS n FROM partidos_presidencial_2026")
                out['partidos'] = cur.fetchone()['n'] or 0

                # ---- Top departamentos por avance escrutinio + preconteo ----
                expr_escr = expr if ultimo > 0 else '0'
                cur.execute(f"""
                    WITH prec_por_depto AS (
                        SELECT coddepto,
                               COUNT(DISTINCT (codmipio, codzona, codpuesto, mesa)) AS mesas_prec
                        FROM preconteo_presidencial_2026
                        WHERE votos > 0
                        GROUP BY coddepto
                    )
                    SELECT d.coddepto, MAX(d.nomdepto) AS nomdepto,
                           COUNT(DISTINCT dm.idmesa) AS mesas_total,
                           COUNT(DISTINCT CASE WHEN {expr_escr} > 0 THEN dm.idmesa END) AS mesas_escr,
                           COALESCE(MAX(pp.mesas_prec), 0) AS mesas_prec
                    FROM divipol_presidencial_2026 d
                    JOIN divipolmesa_presidencial_2026 dm ON dm.iddivipol = d.iddivipol
                    LEFT JOIN seguimiento_escrutinio_presidencial_2026 s ON s.idmesa = dm.idmesa
                    LEFT JOIN prec_por_depto pp ON pp.coddepto = d.coddepto
                    WHERE d.clase='P'
                    GROUP BY d.coddepto
                    ORDER BY d.coddepto
                """)
                out['top_deptos'] = cur.fetchall()

                # ---- Top candidatos por votos (último día escrutinio o preconteo) ----
                if ultimo > 0:
                    cur.execute(f"""
                        SELECT s.codcandidato,
                               (SELECT nomcandidato FROM candidatos_presidencial_2026 WHERE codcandidato=s.codcandidato LIMIT 1) AS nomcandidato,
                               (SELECT nompartido FROM partidos_presidencial_2026 WHERE codpartido=s.codpartido LIMIT 1) AS nompartido,
                               SUM({expr}) AS votos
                        FROM seguimiento_escrutinio_presidencial_2026 s
                        WHERE s.codcandidato NOT IN (996,997,998)
                        GROUP BY s.codcandidato, s.codpartido
                        ORDER BY votos DESC NULLS LAST LIMIT 15
                    """)
                    out['top_candidatos'] = cur.fetchall()
                    out['fuente_top'] = f'Escrutinio día {ultimo}'
                else:
                    cur.execute("""
                        SELECT p.codcandidato,
                               (SELECT nomcandidato FROM candidatos_presidencial_2026 WHERE codcandidato=p.codcandidato LIMIT 1) AS nomcandidato,
                               (SELECT nompartido FROM partidos_presidencial_2026 WHERE codpartido=p.codpartido LIMIT 1) AS nompartido,
                               SUM(p.votos) AS votos
                        FROM preconteo_presidencial_2026 p
                        WHERE p.codcandidato NOT IN (996,997,998)
                        GROUP BY p.codcandidato, p.codpartido
                        ORDER BY votos DESC NULLS LAST LIMIT 15
                    """)
                    out['top_candidatos'] = cur.fetchall()
                    out['fuente_top'] = 'Preconteo'

                # ---- Top candidatos SIEMPRE por preconteo (separado) ----
                cur.execute("""
                    SELECT p.codcandidato,
                           (SELECT nomcandidato FROM candidatos_presidencial_2026 WHERE codcandidato=p.codcandidato LIMIT 1) AS nomcandidato,
                           (SELECT nompartido FROM partidos_presidencial_2026 WHERE codpartido=p.codpartido LIMIT 1) AS nompartido,
                           SUM(p.votos) AS votos
                    FROM preconteo_presidencial_2026 p
                    WHERE p.codcandidato NOT IN (996,997,998)
                    GROUP BY p.codcandidato, p.codpartido
                    ORDER BY votos DESC NULLS LAST LIMIT 15
                """)
                out['top_candidatos_preconteo'] = cur.fetchall()

                # Anexar URL de foto si existe
                _foto_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                         'public', 'static', 'candidatos')
                def _foto_url(cod):
                    if not cod:
                        return None
                    for ext in ('png', 'jpg', 'jpeg'):
                        if os.path.isfile(os.path.join(_foto_dir, f'{cod}.{ext}')):
                            return f'/static/candidatos/{cod}.{ext}'
                    return None
                for c in out['top_candidatos']:
                    c['foto_url'] = _foto_url(c.get('codcandidato'))
                for c in out['top_candidatos_preconteo']:
                    c['foto_url'] = _foto_url(c.get('codcandidato'))

                # ---- Votos especiales totales (último día) ----
                if ultimo > 0:
                    cur.execute(f"""SELECT
                                      SUM(CASE WHEN codcandidato=996 THEN {expr} ELSE 0 END) AS blanco,
                                      SUM(CASE WHEN codcandidato=997 THEN {expr} ELSE 0 END) AS nulo,
                                      SUM(CASE WHEN codcandidato=998 THEN {expr} ELSE 0 END) AS no_marcado
                                    FROM seguimiento_escrutinio_presidencial_2026""")
                else:
                    cur.execute("""SELECT
                                     SUM(CASE WHEN codcandidato=996 THEN votos ELSE 0 END) AS blanco,
                                     SUM(CASE WHEN codcandidato=997 THEN votos ELSE 0 END) AS nulo,
                                     SUM(CASE WHEN codcandidato=998 THEN votos ELSE 0 END) AS no_marcado
                                   FROM preconteo_presidencial_2026""")
                r = cur.fetchone()
                out['especiales'] = {
                    'blanco': int(r['blanco'] or 0), 'nulo': int(r['nulo'] or 0),
                    'no_marcado': int(r['no_marcado'] or 0)
                }

                # ---- Votos especiales SIEMPRE por preconteo (base del % de la lista de preconteo) ----
                cur.execute("""SELECT
                                 SUM(CASE WHEN codcandidato=996 THEN votos ELSE 0 END) AS blanco,
                                 SUM(CASE WHEN codcandidato=997 THEN votos ELSE 0 END) AS nulo,
                                 SUM(CASE WHEN codcandidato=998 THEN votos ELSE 0 END) AS no_marcado
                               FROM preconteo_presidencial_2026""")
                rp = cur.fetchone()
                out['especiales_preconteo'] = {
                    'blanco': int(rp['blanco'] or 0), 'nulo': int(rp['nulo'] or 0),
                    'no_marcado': int(rp['no_marcado'] or 0)
                }

                # ---- Días de seguimiento (cuántos días procesados) ----
                cur.execute("SELECT COUNT(*) AS n FROM dias_escrutinio_presidencial WHERE procesado=TRUE")
                out['dias_procesados'] = cur.fetchone()['n'] or 0

                # ---- Usuarios activos (sesiones / hardcoded skip) ----
                cur.execute("SELECT COUNT(*) AS n FROM usuarios")
                out['usuarios'] = cur.fetchone()['n'] or 0

        # Calcular porcentajes
        tm = out['total_mesas'] or 1
        out['cobertura_preconteo'] = round(out['mesas_preconteo'] * 100.0 / tm, 2)
        out['cobertura_escrutinio'] = round(out['mesas_escrutinio'] * 100.0 / tm, 2)
        out['cobertura_evidencias'] = round(out['evidencias']['mesas'] * 100.0 / tm, 2)
        out['cobertura_e14'] = round(out['e14']['mesas'] * 100.0 / tm, 2)
        out['cobertura_age'] = round(out['age']['mesas'] * 100.0 / tm, 2)

        return jsonify({'success': True, 'data': out, 'ts': int(_time_mod.time())})
    except Exception as e:
        logger.exception('[dashboard/metricas]')
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== MAIN ====================
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5002))
    if '--waitress' in sys.argv:
        from waitress import serve
        logger.info(f"[main] Waitress en :{port}")
        serve(app, host='0.0.0.0', port=port, threads=32)
    else:
        logger.info(f"[main] Flask dev en :{port}")
        app.run(host='0.0.0.0', port=port, debug=True)
